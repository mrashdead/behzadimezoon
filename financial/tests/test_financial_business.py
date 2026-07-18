import json
import warnings
from io import BytesIO

from django.core.exceptions import ValidationError
from django.db.utils import OperationalError
from django.test import RequestFactory, TestCase
from django.urls import reverse
from openpyxl import load_workbook

import jdatetime

from accounts.models import User
from customers.models import Customer
from financial.models import CancellationRecord, DamageRecord, Guarantee, Transaction, FinancialAccount
from unittest.mock import patch
from financial.services import (
    CancellationService,
    DamageService,
    GuaranteeService,
    PaymentService,
    RefundService,
    ReservationFinancialService,
)
from financial.services.dashboard_service import DashboardService
from financial.services.reconciliation_service import ReconciliationService
from financial.services.transaction_service import TransactionService
from financial import views_operations
from products.models import Dress
from reservations.constants import PaymentMethod, ReservationStatus
from reservations.models import AdditionalFee, Reservation


class FinancialBusinessTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user('admin', password='pw', is_superuser=True, role='SUPER_ADMIN')
        cls.seller1 = User.objects.create_user('seller1', password='pw', role='SELLER')
        cls.seller2 = User.objects.create_user('seller2', password='pw', role='SELLER')

        cls.customer = Customer.objects.create(
            bride_first_name='A',
            bride_last_name='B',
            bride_phone='09120000000',
            ceremony_date=jdatetime.date(1402, 1, 1),
            how_to_know='test',
            allow_contact=True,
        )
        cls.dress = Dress.objects.create(code='D001', daily_rent_price=100000)

    def create_reservation(self, **kwargs):
        defaults = {
            'customer': self.customer,
            'dress': self.dress,
            'start_date': jdatetime.date(1402, 1, 1),
            'rental_days': 3,
            'status': ReservationStatus.CONFIRMED,
            'rent_price': self.dress.daily_rent_price,
            'deposit_amount': 0,
            'discount_type': 'NONE',
            'discount_value': 0,
            'discount_amount': 0,
            'final_price': 100000,
            'remaining_amount': 100000,
            'remaining_payment_amount': 0,
            'refunded_amount': 0,
            'payment_method': 'CASH',
            'payment_tracking_code': 'TXN1',
            'guarantee1_type': 'CASH',
            'guarantee1_tracking_code': 'G1',
            'created_by': self.seller1,
        }
        defaults.update(kwargs)
        return Reservation.objects.create(**defaults)

    def test_financial_context_applies_date_and_seller_filters(self):
        reservation_in_range = self.create_reservation(
            start_date=jdatetime.date(1402, 2, 1),
            created_by=self.seller2,
            final_price=120000,
            remaining_amount=120000,
        )
        self.create_reservation(
            start_date=jdatetime.date(1402, 1, 1),
            created_by=self.seller1,
            final_price=90000,
            remaining_amount=90000,
        )

        other_reservation = Reservation.objects.get(created_by=self.seller1, start_date=jdatetime.date(1402, 1, 1))

        context = DashboardService.get_financial_context(filters={
            'date_from': '1402/01/15',
            'date_to': '1402/02/15',
            'seller_id': str(self.seller2.pk),
        })

        recent_ids = [item.pk for item in context['recent_reservations']]
        self.assertIn(reservation_in_range.pk, recent_ids)
        self.assertNotIn(other_reservation.pk, recent_ids)

    def test_transaction_date_filters_do_not_emit_naive_datetime_warnings(self):
        self.create_reservation(
            start_date=jdatetime.date(1402, 2, 1),
            created_by=self.seller2,
            final_price=120000,
            remaining_amount=120000,
        )

        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter('always')
            DashboardService.get_financial_context(filters={
                'date_from': '1402/01/15',
                'date_to': '1402/02/15',
                'seller_id': str(self.seller2.pk),
            })

        self.assertFalse(any('received a naive datetime' in str(w.message) for w in caught))

    def test_financial_export_matches_page_results_for_selected_date_range(self):
        in_range = self.create_reservation(
            start_date=jdatetime.date(1402, 2, 1),
            created_by=self.seller2,
            final_price=120000,
            remaining_amount=120000,
        )
        self.create_reservation(
            start_date=jdatetime.date(1402, 1, 1),
            created_by=self.seller1,
            final_price=90000,
            remaining_amount=90000,
        )

        filters = {
            'date_from': '1402/01/15',
            'date_to': '1402/02/15',
            'seller_id': str(self.seller2.pk),
        }

        self.client.login(username='seller1', password='pw')
        page_context = DashboardService.get_financial_context(filters=filters)
        response = self.client.get(reverse('financial:export_excel'), data=filters)

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content), read_only=True, data_only=True)
        worksheet = workbook.active
        exported_ids = [row[0] for row in worksheet.iter_rows(min_row=2, values_only=True) if row[0] is not None]

        page_ids = [item.pk for item in page_context['recent_reservations']]
        self.assertEqual(page_ids, exported_ids)
        self.assertIn(in_range.pk, page_ids)

    def test_financial_export_without_filters_returns_all_reservations(self):
        reservation1 = self.create_reservation(
            start_date=jdatetime.date(1402, 3, 1),
            final_price=100000,
            remaining_amount=100000,
        )
        reservation2 = self.create_reservation(
            start_date=jdatetime.date(1402, 3, 10),
            final_price=150000,
            remaining_amount=150000,
        )

        self.client.login(username='seller1', password='pw')
        response = self.client.get(reverse('financial:export_excel'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content), read_only=True, data_only=True)
        worksheet = workbook.active
        exported_ids = [row[0] for row in worksheet.iter_rows(min_row=2, values_only=True) if row[0] is not None]

        total_reservations = Reservation.objects.filter(is_deleted=False).count()
        self.assertEqual(len(exported_ids), total_reservations)
        self.assertIn(reservation1.pk, exported_ids)
        self.assertIn(reservation2.pk, exported_ids)

    def test_default_cash_account_is_reused_when_already_present(self):
        FinancialAccount.objects.create(
            code='CASH_DEFAULT',
            name='Existing Cash Account',
            account_type=FinancialAccount.AccountType.CASH,
            balance=0,
            description='existing',
            is_active=True,
        )
        reservation = self.create_reservation(deposit_amount=0, remaining_payment_amount=0, final_price=100000, remaining_amount=100000)

        tx = TransactionService.create(
            reservation=reservation,
            transaction_type=Transaction.TransactionType.DEPOSIT,
            amount=1000,
            created_by=self.admin,
        )

        self.assertEqual(tx.account.code, 'CASH_DEFAULT')
        self.assertEqual(FinancialAccount.objects.filter(code='CASH_DEFAULT').count(), 1)

    def test_deposit_and_final_settlement_updates_reservation_balance(self):
        reservation = self.create_reservation(deposit_amount=0, remaining_payment_amount=0, final_price=100000, remaining_amount=100000)

        PaymentService.record_deposit(
            reservation=reservation,
            amount=50000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Initial deposit',
        )
        reservation.refresh_from_db()
        self.assertEqual(reservation.deposit_amount, 50000)
        self.assertEqual(reservation.financial_snapshot['event_type'], 'deposit')
        self.assertEqual(reservation.financial_snapshot['deposit_amount'], 50000)

        PaymentService.record_balance_payment(
            reservation=reservation,
            amount=50000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Final settlement',
        )
        reservation.refresh_from_db()

        self.assertEqual(reservation.remaining_payment_amount, 50000)
        self.assertEqual(reservation.remaining_amount, 0)
        self.assertEqual(reservation.net_cash_inflow, 100000)
        self.assertEqual(reservation.financial_snapshot['event_type'], 'balance_payment')
        self.assertEqual(TransactionService.reservation_cash_inflow(reservation), 100000)

    def test_discount_calculation_and_invalid_discount_rejected(self):
        reservation = self.create_reservation()
        reservation.discount_amount = 20000
        ReservationFinancialService.calculate_final_price(reservation)
        self.assertEqual(reservation.final_price, 80000)

        reservation.discount_amount = 120000
        with self.assertRaises(ValidationError):
            ReservationFinancialService.validate_discount(reservation)

        reservation.discount_amount = -100
        with self.assertRaises(ValidationError):
            ReservationFinancialService.validate_discount(reservation)

    def test_partial_payment_status_after_discount_and_deposit(self):
        custom_dress = Dress.objects.create(code='D002', daily_rent_price=35000)
        reservation = self.create_reservation(
            dress=custom_dress,
            rent_price=35000,
            discount_type='AMOUNT',
            discount_value=7000,
            discount_amount=7000,
            final_price=28000,
            deposit_amount=20000,
            remaining_payment_amount=0,
            remaining_amount=8000,
        )

        ReservationFinancialService.update_financial_status(reservation)

        self.assertEqual(reservation.remaining_amount, 8000)
        self.assertEqual(reservation.payment_status, Reservation.PAYMENT_PARTIAL)

    def test_dashboard_totals_include_additional_fee_revenue(self):
        reservation = self.create_reservation(final_price=100000, remaining_amount=100000)
        AdditionalFee.objects.create(reservation=reservation, title='اتوکشی', amount=15000, created_by=self.admin)
        AdditionalFee.objects.create(reservation=reservation, title='ارسال', amount=5000, created_by=self.admin)

        context = DashboardService.get_financial_context(filters={})

        self.assertEqual(context['totals']['total_additional_fee_revenue'], 20000)

    def test_excel_export_contains_only_requested_report_fields(self):
        self.create_reservation(
            final_price=100000,
            remaining_amount=100000,
            payment_status=Reservation.PAYMENT_PARTIAL,
            payment_method='CASH',
            payment_tracking_code='TK-1',
            remaining_payment_method=PaymentMethod.POS,
            remaining_payment_tracking_code='TK-2',
            remaining_paid_at=jdatetime.datetime(1402, 1, 2, 10, 0),
            rent_price=100000,
            discount_type='AMOUNT',
            discount_value=10000,
            discount_amount=10000,
            deposit_amount=50000,
            refunded_amount=5000,
            cancellation_fee=7000,
        )

        self.client.login(username='seller1', password='pw')
        response = self.client.get(reverse('financial:export_excel'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content), read_only=True, data_only=True)
        worksheet = workbook['رزروها']
        headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

        self.assertIn('شناسه رزرو', headers)
        self.assertIn('تخفیف', headers)
        self.assertIn('مبلغ نهایی', headers)
        self.assertIn('بیعانه', headers)
        self.assertIn('باقی‌مانده', headers)
        self.assertIn('هزینه لغو', headers)
        self.assertNotIn('نام عروس', headers)

    def test_excel_export_creates_summary_and_transaction_sheets_for_accounting_variant(self):
        self.create_reservation(final_price=100000, remaining_amount=100000, deposit_amount=50000)
        self.client.login(username='seller1', password='pw')
        response = self.client.get(reverse('financial:export_excel'), data={'variant': 'accounting'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content), read_only=True, data_only=True)

        self.assertEqual(set(workbook.sheetnames), {'خلاصه', 'رزروها', 'تراکنش‌ها', 'بازرسی مهاجرت'})
        self.assertIn('خلاصه', workbook.sheetnames)
        self.assertIn('تراکنش‌ها', workbook.sheetnames)
        self.assertIn('بازرسی مهاجرت', workbook.sheetnames)

    def test_excel_export_summary_sheet_has_executive_layout(self):
        self.create_reservation(final_price=100000, remaining_amount=100000, deposit_amount=50000)
        self.client.login(username='seller1', password='pw')
        response = self.client.get(reverse('financial:export_excel'), data={'variant': 'accounting'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content), read_only=True, data_only=True)
        summary_sheet = workbook['خلاصه']

        self.assertEqual(summary_sheet['A1'].value, 'خلاصه مدیریتی')
        self.assertEqual(summary_sheet['A3'].value, 'متریک')
        self.assertEqual(summary_sheet['B3'].value, 'ارزش')

    def test_excel_export_translates_values_and_headers_to_persian(self):
        self.create_reservation(
            final_price=100000,
            remaining_amount=100000,
            payment_status=Reservation.PAYMENT_PAID,
            payment_method='CASH',
            status=ReservationStatus.CONFIRMED,
        )
        self.client.login(username='seller1', password='pw')
        response = self.client.get(reverse('financial:export_excel'), data={'variant': 'accounting'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content), read_only=True, data_only=True)

        reservation_headers = [cell.value for cell in next(workbook['رزروها'].iter_rows(min_row=1, max_row=1))]
        self.assertEqual(reservation_headers.count('وضعیت'), 1)
        self.assertNotIn('Unnamed: 1', reservation_headers)

        audit_headers = [cell.value for cell in next(workbook['بازرسی مهاجرت'].iter_rows(min_row=1, max_row=1))]
        self.assertIn('شناسه', audit_headers)
        self.assertIn('تاریخ ایجاد', audit_headers)
        self.assertIn('شناسه تراکنش', audit_headers)
        self.assertNotIn('Unnamed: 1', audit_headers)

        status_idx = reservation_headers.index('وضعیت')
        payment_status_idx = reservation_headers.index('وضعیت پرداخت')
        payment_method_idx = reservation_headers.index('روش پرداخت')
        first_row = next(workbook['رزروها'].iter_rows(min_row=2, max_row=2))
        self.assertEqual(first_row[status_idx].value, 'تایید شده')
        self.assertEqual(first_row[payment_status_idx].value, 'پرداخت کامل')
        self.assertEqual(first_row[payment_method_idx].value, 'نقدی')

    def test_penalty_payments_are_recorded_and_counted_as_revenue(self):
        reservation = self.create_reservation(final_price=100000, remaining_amount=100000)
        reservation.cancellation_fee = 20000
        reservation.damage_amount = 30000
        reservation.save(update_fields=['cancellation_fee', 'damage_amount'])

        PaymentService.record_penalty_payment(
            reservation=reservation,
            amount=20000,
            penalty_type='CANCELLATION',
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            external_reference='PEN-1',
            note='Cancellation penalty received',
        )
        PaymentService.record_penalty_payment(
            reservation=reservation,
            amount=30000,
            penalty_type='DAMAGE',
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            external_reference='PEN-2',
            note='Damage penalty received',
        )

        totals = TransactionService.aggregate_reservation_totals(reservation)
        context = DashboardService.get_financial_context(filters={})

        self.assertEqual(totals.get('total_penalty_income', 0), 50000)
        self.assertEqual(context['totals']['total_revenue'], 50000)
        self.assertEqual(context['totals']['total_cash_inflow'], 50000)

    def test_seller_can_record_penalty_payment_for_their_reservation(self):
        reservation = self.create_reservation(
            created_by=self.seller1,
            final_price=100000,
            remaining_amount=100000,
            cancellation_fee=15000,
        )

        self.client.login(username='seller1', password='pw')
        response = self.client.post(
            reverse('reservations:penalty_payment', args=[reservation.pk]),
            {
                'penalty_type': 'CANCELLATION',
                'penalty_amount': '15000',
                'penalty_payment_method': PaymentMethod.CASH,
                'penalty_payment_tracking_code': 'PEN-SELLER',
            },
            content_type='application/x-www-form-urlencoded',
        )

        reservation.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        self.assertEqual(reservation.cancellation_fee_paid_amount, 15000)

    def test_penalty_payment_accepts_cash_payment_without_tracking_code(self):
        reservation = self.create_reservation(
            created_by=self.seller1,
            final_price=100000,
            remaining_amount=100000,
            cancellation_fee=15000,
            status=ReservationStatus.CANCELLED,
        )

        self.client.login(username='seller1', password='pw')
        response = self.client.post(
            reverse('reservations:penalty_payment', args=[reservation.pk]),
            {
                'penalty_type': 'CANCELLATION',
                'penalty_amount': '15000',
                'penalty_payment_method': PaymentMethod.CASH,
                'penalty_payment_tracking_code': '',
            },
            content_type='application/x-www-form-urlencoded',
        )

        reservation.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        self.assertEqual(reservation.cancellation_fee_paid_amount, 15000)

    def test_penalty_payment_uses_cancellation_record_when_reservation_fee_field_is_empty(self):
        reservation = self.create_reservation(
            final_price=100000,
            remaining_amount=100000,
            status=ReservationStatus.CANCELLED,
        )
        CancellationRecord.objects.create(
            reservation=reservation,
            reason='Customer cancelled',
            cancelled_by=self.admin,
            penalty_amount=15000,
            refund_amount=0,
        )
        reservation.cancellation_fee = None
        reservation.cancellation_fee_paid_amount = 0
        reservation.save(update_fields=['cancellation_fee', 'cancellation_fee_paid_amount'])

        state = PaymentService.get_penalty_payment_state(reservation, 'CANCELLATION')

        self.assertTrue(state['is_allowed'])
        self.assertEqual(state['remaining_amount'], 15000)

    def test_penalty_payment_requires_a_recorded_cancellation_penalty(self):
        reservation = self.create_reservation(
            final_price=100000,
            remaining_amount=100000,
            status=ReservationStatus.CONFIRMED,
        )

        with self.assertRaises(ValidationError):
            PaymentService.record_penalty_payment(
                reservation=reservation,
                amount=15000,
                penalty_type='CANCELLATION',
                created_by=self.admin,
                payment_method=PaymentMethod.CASH,
                external_reference='PEN-INVALID-STATUS',
                note='Invalid penalty test',
            )

    def test_penalty_payment_rejects_duplicate_payment_when_full_amount_is_already_paid(self):
        reservation = self.create_reservation(
            final_price=100000,
            remaining_amount=100000,
            cancellation_fee=15000,
            cancellation_fee_paid_amount=15000,
            status=ReservationStatus.CANCELLED,
        )

        with self.assertRaises(ValidationError):
            PaymentService.record_penalty_payment(
                reservation=reservation,
                amount=15000,
                penalty_type='CANCELLATION',
                created_by=self.admin,
                payment_method=PaymentMethod.CASH,
                external_reference='PEN-DUPLICATE',
                note='Duplicate payment test',
            )

    def test_penalty_payment_records_transaction_and_updates_paid_amounts(self):
        reservation = self.create_reservation(
            final_price=100000,
            remaining_amount=100000,
            cancellation_fee=15000,
            status=ReservationStatus.CANCELLED,
        )

        tx = PaymentService.record_penalty_payment(
            reservation=reservation,
            amount=15000,
            penalty_type='CANCELLATION',
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            external_reference='PEN-OK',
            note='Valid payment test',
        )

        reservation.refresh_from_db()
        self.assertEqual(tx.transaction_type, Transaction.TransactionType.PENALTY_INCOME)
        self.assertEqual(reservation.cancellation_fee_paid_amount, 15000)
        self.assertEqual(Transaction.objects.filter(reservation=reservation, transaction_type=Transaction.TransactionType.PENALTY_INCOME).count(), 1)

    def test_penalty_payment_retries_after_transient_database_lock(self):
        reservation = self.create_reservation(
            final_price=100000,
            remaining_amount=100000,
            cancellation_fee=15000,
            status=ReservationStatus.CANCELLED,
        )

        original_create_transaction = TransactionService.create_transaction
        attempts = {'count': 0}

        def flaky_create_transaction(*args, **kwargs):
            attempts['count'] += 1
            if attempts['count'] == 1:
                raise OperationalError('database is locked')
            return original_create_transaction(*args, **kwargs)

        with patch('financial.services.payment_service.TransactionService.create_transaction', side_effect=flaky_create_transaction):
            tx = PaymentService.record_penalty_payment(
                reservation=reservation,
                amount=15000,
                penalty_type='CANCELLATION',
                created_by=self.admin,
                payment_method=PaymentMethod.CASH,
                external_reference='PEN-RETRY',
                note='Retry test',
            )

        reservation.refresh_from_db()
        self.assertEqual(attempts['count'], 2)
        self.assertEqual(reservation.cancellation_fee_paid_amount, 15000)
        self.assertEqual(tx.amount, 15000)

    def test_financial_dashboard_includes_cancelled_reservations_in_reservation_list(self):
        cancelled_reservation = self.create_reservation(
            status=ReservationStatus.CANCELLED,
            final_price=80000,
            remaining_amount=80000,
            cancellation_fee=20000,
        )
        self.create_reservation(
            status=ReservationStatus.DELIVERED,
            final_price=90000,
            remaining_amount=0,
        )

        context = DashboardService.get_financial_context(filters={})
        recent_ids = [item.pk for item in context['recent_reservations']]

        self.assertIn(cancelled_reservation.pk, recent_ids)
        self.assertEqual(len(recent_ids), 2)

    def test_refund_view_returns_json_validation_error_when_no_refund_amount_is_allowed(self):
        reservation = self.create_reservation(final_price=100000, remaining_amount=100000)
        factory = RequestFactory()
        request = factory.post(
            reverse('financial:create_transaction', args=[reservation.pk]),
            {
                'type': 'REFUND',
                'amount': '10000',
                'payment_method': PaymentMethod.CASH,
                'external_reference': 'REF-1',
                'note': 'refund attempt',
            },
        )
        request.user = self.admin

        response = views_operations.create_transaction_view(request, reservation.pk)

        self.assertEqual(response.status_code, 400)
        payload = json.loads(response.content)
        self.assertIn('مبلغ بازپرداخت', payload['error'])

    def test_reservation_financial_view_includes_additional_fee_breakdown(self):
        reservation = self.create_reservation(final_price=100000, remaining_amount=100000)
        AdditionalFee.objects.create(reservation=reservation, title='هزینه اضافی', amount=25000, created_by=self.admin)

        self.client.login(username='admin', password='pw')
        response = self.client.get(reverse('financial:reservation_financial', args=[reservation.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'هزینه‌های جانبی')
        self.assertContains(response, 'هزینه اضافی')
        self.assertContains(response, '25,000')

    def test_multiple_payments_and_refund_validation(self):
        reservation = self.create_reservation(deposit_amount=0, remaining_payment_amount=0, final_price=100000, remaining_amount=100000)

        PaymentService.record_deposit(
            reservation=reservation,
            amount=20000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Deposit',
        )
        PaymentService.record_balance_payment(
            reservation=reservation,
            amount=50000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='First payment',
        )
        PaymentService.record_balance_payment(
            reservation=reservation,
            amount=40000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Overpayment',
        )
        reservation.refresh_from_db()

        self.assertEqual(reservation.remaining_amount, 0)
        self.assertEqual(reservation.remaining_payment_amount, 90000)
        self.assertEqual(ReservationFinancialService.allowable_refund_amount(reservation), 10000)

        with self.assertRaises(ValidationError):
            RefundService.record_refund(
                reservation=reservation,
                amount=20000,
                created_by=self.admin,
                payment_method=PaymentMethod.CASH,
                note='Too much refund',
            )

        refund_tx = RefundService.record_refund(
            reservation=reservation,
            amount=10000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Refund overpayment',
        )
        self.assertEqual(refund_tx.type, Transaction.Type.REFUND)
        reservation.refresh_from_db()
        self.assertEqual(reservation.refunded_amount, 10000)

    def test_cancellation_with_full_refund_generates_refund_transaction(self):
        reservation = self.create_reservation(deposit_amount=0, remaining_payment_amount=0)
        PaymentService.record_deposit(
            reservation=reservation,
            amount=100000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Deposit for cancellation',
        )
        PaymentService.record_balance_payment(
            reservation=reservation,
            amount=10000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Overpayment before cancellation',
        )
        reservation.refresh_from_db()

        cancellation = CancellationService.create_cancellation_record(
            reservation=reservation,
            reason='Customer cancelled',
            created_by=self.admin,
            refund_amount=10000,
            penalty_amount=0,
            payment_method=PaymentMethod.CASH,
            note='Full refund',
        )

        self.assertIsInstance(cancellation, CancellationRecord)
        self.assertEqual(cancellation.refund_amount, 10000)
        self.assertEqual(cancellation.penalty_amount, 0)
        self.assertIsNotNone(cancellation.related_transaction)
        self.assertEqual(cancellation.related_transaction.type, Transaction.Type.REFUND)

    def test_cancellation_partial_refund_and_penalty_creates_both_transactions(self):
        reservation = self.create_reservation(deposit_amount=0, remaining_payment_amount=0)
        PaymentService.record_deposit(
            reservation=reservation,
            amount=100000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Deposit for partial cancellation',
        )
        PaymentService.record_balance_payment(
            reservation=reservation,
            amount=10000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Overpayment before partial cancellation',
        )
        reservation.refresh_from_db()

        cancellation = CancellationService.create_cancellation_record(
            reservation=reservation,
            reason='Partial refund on cancellation',
            created_by=self.admin,
            refund_amount=5000,
            penalty_amount=2000,
            payment_method=PaymentMethod.CASH,
            note='Partial refund',
        )

        self.assertEqual(cancellation.refund_amount, 5000)
        self.assertEqual(cancellation.penalty_amount, 2000)
        self.assertEqual(Transaction.objects.filter(type=Transaction.Type.REFUND, reservation=reservation).count(), 1)
        self.assertEqual(Transaction.objects.filter(type=Transaction.Type.CANCELLATION_FEE, reservation=reservation).count(), 1)

    def test_cancellation_penalty_is_saved_on_reservation(self):
        reservation = self.create_reservation(deposit_amount=0, remaining_payment_amount=0)

        CancellationService.create_cancellation_record(
            reservation=reservation,
            reason='Customer cancelled',
            created_by=self.admin,
            refund_amount=0,
            penalty_amount=15000,
            payment_method=PaymentMethod.CASH,
            note='Penalty only cancellation',
        )

        reservation.refresh_from_db()
        self.assertEqual(reservation.cancellation_fee, 15000)

    def test_damage_registration_and_payment_links_consistently(self):
        reservation = self.create_reservation(deposit_amount=5000, remaining_payment_amount=0, final_price=5000, remaining_amount=0)

        damage_record = DamageService.record_damage(
            reservation=reservation,
            customer=self.customer,
            damage_type='Tear',
            amount=20000,
            description='Minor tear',
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            external_reference='D1',
        )
        self.assertIsInstance(damage_record, DamageRecord)
        self.assertFalse(damage_record.collected)
        self.assertEqual(damage_record.related_transaction.type, Transaction.Type.DAMAGE_CHARGE)
        self.assertEqual(damage_record.related_transaction.reservation_id, reservation.pk)

        payment_tx = DamageService.record_damage_payment(
            damage_record=damage_record,
            amount=20000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            external_reference='D1PAY',
            note='Damage payment',
        )
        damage_record.refresh_from_db()
        self.assertTrue(damage_record.collected)
        self.assertEqual(payment_tx.type, Transaction.Type.DAMAGE_PAYMENT)
        self.assertEqual(damage_record.related_transaction_id, payment_tx.pk)

    def test_guarantee_registration_and_return_flow(self):
        reservation = self.create_reservation()
        guarantee = GuaranteeService.create_guarantee(
            reservation=reservation,
            customer=self.customer,
            tracking_code='G-100',
            guarantee_type='CASH',
            estimated_value=50000,
            notes='Guarantee created',
        )
        self.assertIsInstance(guarantee, Guarantee)
        self.assertEqual(guarantee.status, Guarantee.RECEIVED)

        duplicate = GuaranteeService.create_guarantee(
            reservation=reservation,
            customer=self.customer,
            tracking_code='G-100',
            guarantee_type='CASH',
        )
        self.assertEqual(Guarantee.objects.filter(tracking_code='G-100').count(), 2)

        self.client.login(username='admin', password='pw')
        response = self.client.post(reverse('financial:return_guarantee', args=[guarantee.pk]))
        self.assertEqual(response.status_code, 200)
        guarantee.refresh_from_db()
        self.assertEqual(guarantee.status, Guarantee.RETURNED)

    def test_reconciliation_detects_mismatch_between_snapshot_and_ledger(self):
        reservation = self.create_reservation(deposit_amount=50000, remaining_payment_amount=100000, final_price=100000, remaining_amount=0)
        TransactionService.create_deposit(
            reservation=reservation,
            amount=50000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Deposit',
        )
        TransactionService.create_final_payment(
            reservation=reservation,
            amount=50000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Partial settlement',
        )

        discrepancy = ReconciliationService.reservation_discrepancies(reservation)
        self.assertTrue(discrepancy['has_discrepancy'])
        self.assertLess(discrepancy['cash_difference'], 0)
        self.assertEqual(discrepancy['suggested_action'], 'adjustment')
        self.assertEqual(discrepancy['expected_cash'], 100000)

    def test_adjustment_transaction_resolves_reconciliation_discrepancy(self):
        reservation = self.create_reservation(deposit_amount=50000, remaining_payment_amount=100000, final_price=100000, remaining_amount=0)
        TransactionService.create_deposit(
            reservation=reservation,
            amount=50000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Deposit',
        )
        TransactionService.create_final_payment(
            reservation=reservation,
            amount=50000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Final payment',
        )

        discrepancy = ReconciliationService.reservation_discrepancies(reservation)
        self.assertTrue(discrepancy['has_discrepancy'])
        self.assertEqual(discrepancy['suggested_action'], 'adjustment')

        TransactionService.create(
            reservation=reservation,
            transaction_type=Transaction.Type.ADJUSTMENT,
            amount=50000,
            created_by=self.admin,
            note='Reconciliation adjustment',
        )

        discrepancy_after = ReconciliationService.reservation_discrepancies(reservation)
        self.assertFalse(discrepancy_after['has_discrepancy'])
        self.assertEqual(discrepancy_after['cash_difference'], 0)

    def test_financial_snapshot_updates_on_each_transaction(self):
        reservation = self.create_reservation(deposit_amount=0, remaining_payment_amount=0, final_price=100000, remaining_amount=100000)

        PaymentService.record_deposit(
            reservation=reservation,
            amount=40000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Deposit snapshot',
        )
        reservation.refresh_from_db()
        self.assertEqual(reservation.financial_snapshot['event_type'], 'deposit')
        self.assertEqual(reservation.total_cash_collected_snapshot, 40000)

        PaymentService.record_balance_payment(
            reservation=reservation,
            amount=60000,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='Payment snapshot',
        )
        reservation.refresh_from_db()
        self.assertEqual(reservation.financial_snapshot['event_type'], 'balance_payment')

        RefundService.record_refund(
            reservation=reservation,
            amount=0,
            created_by=self.admin,
            payment_method=PaymentMethod.CASH,
            note='No refund',
        )
        reservation.refresh_from_db()
        self.assertIn('final_price', reservation.financial_snapshot)

    def test_operational_views_reject_invalid_inputs_and_access(self):
        reservation = self.create_reservation(created_by=self.seller1)
        self.client.login(username='seller2', password='pw')

        response = self.client.post(reverse('financial:add_guarantee', args=[reservation.pk]), {
            'tracking_code': 'X1',
            'guarantee_type': 'CASH',
        })
        self.assertEqual(response.status_code, 403)

        self.client.login(username='admin', password='pw')
        response = self.client.post(reverse('financial:create_transaction', args=[reservation.pk]), {
            'type': 'DEPOSIT',
            'amount': -100,
            'payment_method': PaymentMethod.CASH,
        })
        self.assertEqual(response.status_code, 400)
        self.assertIn('amount', response.json()['error'])

        response = self.client.post(reverse('financial:add_damage', args=[reservation.pk]), {
            'damage_type': 'Leak',
            'amount': -5000,
        })
        self.assertEqual(response.status_code, 400)
        self.assertIn('amount', response.json()['error'])

        response = self.client.post(reverse('financial:cancel_reservation', args=[reservation.pk]), {
            'refund_amount': -1000,
            'penalty_amount': 0,
        })
        self.assertEqual(response.status_code, 400)
        self.assertIn('refund_amount', response.json()['error'])

    def test_reservation_financial_view_renders_summary(self):
        reservation = self.create_reservation()
        self.client.login(username='admin', password='pw')
        response = self.client.get(reverse('financial:reservation_financial', args=[reservation.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'بهای پایه')

    def test_reservation_financial_view_shows_cancellation_penalty_payment_button_for_cancelled_reservation_with_cancellation_record(self):
        reservation = self.create_reservation(status=ReservationStatus.CONFIRMED)
        CancellationService.create_cancellation_record(
            reservation=reservation,
            reason='Customer cancelled',
            created_by=self.admin,
            refund_amount=0,
            penalty_amount=15000,
            payment_method=PaymentMethod.CASH,
            note='Penalty on cancellation',
        )
        reservation.refresh_from_db()

        self.client.login(username='admin', password='pw')
        response = self.client.get(reverse('financial:reservation_financial', args=[reservation.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'پرداخت مانده جریمه لغو')
