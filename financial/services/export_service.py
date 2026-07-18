import datetime
from io import BytesIO

from django.db.models import Q, Sum, F
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from financial.models import Transaction, FinancialAccount, TransactionCategory
from reservations.constants import PaymentMethod, ReservationStatus
from reservations.models import AdditionalFee, Reservation
from reservations.utils import parse_reservation_date


class FinancialExportService:
    DEFAULT_VARIANT = 'management'
    SUPPORTED_VARIANTS = {'management', 'accounting', 'migration'}

    SHEET_DEFINITIONS = {
        'Summary': {'title': 'خلاصه', 'required': True},
        'Reservations': {'title': 'رزروها', 'required': True},
        'Transactions': {'title': 'تراکنش‌ها', 'required': False},
        'Audit_Migration': {'title': 'بازرسی مهاجرت', 'required': False},
    }

    VALUE_TRANSLATIONS = {
        'CONFIRMED': 'تایید شده',
        'CANCELLED': 'لغو شده',
        'DELIVERED': 'تحویل شده',
        'PARTIAL': 'پرداخت ناقص',
        'PAID': 'پرداخت کامل',
        'UNPAID': 'پرداخت نشده',
        'REFUNDED': 'بازپرداخت شده',
        'No': 'خیر',
        'Yes': 'بله',
        'CASH': 'نقدی',
        'TRANSFER': 'انتقال بانکی',
        'PAYMENT': 'پرداخت',
        'DEPOSIT': 'بیعانه',
        'REFUND': 'بازپرداخت',
        'CANCELLATION_FEE': 'هزینه لغو',
        'DAMAGE_CHARGE': 'هزینه خسارت',
        'ACCRUAL': 'تعهدی',
        'UNKNOWN': 'ناشناخته',
    }

    @staticmethod
    def normalize_variant(variant):
        if not variant:
            return FinancialExportService.DEFAULT_VARIANT
        value = str(variant).strip().lower()
        if value not in FinancialExportService.SUPPORTED_VARIANTS:
            return FinancialExportService.DEFAULT_VARIANT
        return value

    @staticmethod
    def _normalize_date_value(value):
        if value is None:
            return ''
        if isinstance(value, datetime.datetime):
            try:
                return value.date().isoformat()
            except Exception:
                return value.isoformat()
        if isinstance(value, datetime.date):
            return value.isoformat()
        return str(value)

    @staticmethod
    def _normalize_datetime_value(value):
        if value is None:
            return ''
        if isinstance(value, datetime.datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, datetime.date):
            return value.strftime('%Y-%m-%d')
        return str(value)

    @staticmethod
    def _normalize_boolean_value(value):
        if value in (None, ''):
            return 'خیر'
        if isinstance(value, bool):
            return 'بله' if value else 'خیر'
        if isinstance(value, (int, float)):
            return 'بله' if value else 'خیر'
        return FinancialExportService._translate_value(value)

    @staticmethod
    def _translate_value(value):
        if value in (None, ''):
            return ''
        if isinstance(value, bool):
            return 'بله' if value else 'خیر'
        if isinstance(value, (int, float)):
            return value
        text = str(value).strip()
        if not text:
            return ''
        return FinancialExportService.VALUE_TRANSLATIONS.get(text, FinancialExportService.VALUE_TRANSLATIONS.get(text.upper(), text))

    @staticmethod
    def _normalize_money(value):
        if value in (None, ''):
            return 0
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    @staticmethod
    def _resolve_row_value(row, field):
        if field == 'status_code':
            return row.get('status_label', row.get(field, ''))
        if field == 'payment_status_code':
            return row.get('payment_status_label', row.get(field, ''))
        if field == 'payment_method_code':
            return FinancialExportService._translate_value(row.get(field, ''))
        return row.get(field, '')

    @staticmethod
    def _normalize_status_label(value):
        if not value:
            return '-'
        translated = FinancialExportService._translate_value(value)
        if translated != value:
            return translated
        raw_value = dict(ReservationStatus.CHOICES).get(value, value)
        return FinancialExportService._translate_value(raw_value)

    @staticmethod
    def _normalize_payment_status_label(value):
        if not value:
            return '-'
        return FinancialExportService._translate_value(value)

    @staticmethod
    def _user_label(user):
        if not user:
            return '-'
        return getattr(user, 'get_full_name', lambda: str(user))() or str(user)

    @staticmethod
    def _parse_filter_date(value):
        if value is None:
            return None
        if isinstance(value, datetime.datetime):
            return value.date()
        if isinstance(value, datetime.date):
            return value
        if hasattr(value, 'togregorian'):
            try:
                return value.togregorian()
            except Exception:
                return value
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            return parse_reservation_date(value)
        return None

    @staticmethod
    def _build_queryset(filters=None):
        filters = filters or {}
        queryset = Reservation.objects.filter(is_deleted=False)
        seller_id = filters.get('seller_id') or filters.get('seller')
        if seller_id:
            try:
                queryset = queryset.filter(created_by_id=int(seller_id))
            except (TypeError, ValueError):
                pass
        date_from = FinancialExportService._parse_filter_date(filters.get('date_from'))
        date_to = FinancialExportService._parse_filter_date(filters.get('date_to'))
        if date_from:
            queryset = queryset.filter(start_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(start_date__lte=date_to)
        return queryset.select_related('customer', 'dress', 'created_by', 'updated_by', 'archived_by').prefetch_related('transactions', 'additional_fees')

    @staticmethod
    def _build_summary_rows(reservations, variant, filters=None):
        filters = filters or {}
        total_final_amount = sum((r.final_price or 0) for r in reservations)
        total_deposit = sum((r.deposit_amount or 0) for r in reservations)
        total_remaining_balance = sum((r.remaining_amount or 0) for r in reservations)
        total_cancellation_fees = sum((r.cancellation_fee or 0) for r in reservations)
        cancelled_count = sum(1 for r in reservations if getattr(r, 'status', None) == ReservationStatus.CANCELLED)
        damaged_count = sum(1 for r in reservations if getattr(r, 'item_damaged', False) or (getattr(r, 'damage_amount', None) or 0) > 0)
        total_collected = sum((r.deposit_amount or 0) for r in reservations)
        total_uncollected = sum((r.remaining_amount or 0) for r in reservations)
        settlement_ratio = (total_collected / total_final_amount) if total_final_amount else 0

        return [
            ('export_variant', variant),
            ('generated_at', timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')),
            ('date_from', filters.get('date_from') or '-'),
            ('date_to', filters.get('date_to') or '-'),
            ('total_reservations', len(reservations)),
            ('total_final_amount', total_final_amount),
            ('total_deposit', total_deposit),
            ('total_remaining_balance', total_remaining_balance),
            ('total_cancellation_fees', total_cancellation_fees),
            ('cancelled_count', cancelled_count),
            ('damaged_count', damaged_count),
            ('total_collected', total_collected),
            ('total_uncollected', total_uncollected),
            ('settlement_ratio', settlement_ratio),
        ]

    @staticmethod
    def _build_reservation_rows(reservations, variant):
        rows = []
        for reservation in reservations:
            customer = getattr(reservation, 'customer', None)
            dress = getattr(reservation, 'dress', None)
            additional_fee_total = sum((fee.amount or 0) for fee in reservation.additional_fees.filter(is_deleted=False)) if hasattr(reservation, 'additional_fees') else 0
            transaction_total = sum((tx.amount or 0) for tx in reservation.transactions.filter(transaction_status=Transaction.TransactionStatus.POSTED, is_voided=False)) if hasattr(reservation, 'transactions') else 0
            collected = (reservation.deposit_amount or 0) + transaction_total - (reservation.refunded_amount or 0)
            remaining_balance = max((reservation.final_price or 0) + (reservation.cancellation_fee or 0) + (reservation.damage_amount or 0) - collected, 0)
            settled = 'PAID' if remaining_balance == 0 and (reservation.final_price or 0) > 0 else 'PARTIAL' if collected > 0 else 'UNPAID'
            damage_flag = bool(getattr(reservation, 'item_damaged', False) or (getattr(reservation, 'damage_amount', None) or 0) > 0)

            row = {
                'reservation_id': reservation.pk,
                'contract_number': getattr(reservation, 'contract_number', '') or '-',
                'customer_id': getattr(customer, 'pk', None),
                'customer_name': str(customer) if customer else '-',
                'bride_phone': getattr(customer, 'bride_phone', '') or '-',
                'dress_id': getattr(dress, 'pk', None),
                'dress_code': getattr(dress, 'code', '') or '-',
                'start_date': FinancialExportService._normalize_date_value(getattr(reservation, 'start_date', None)),
                'start_date_jalali': FinancialExportService._normalize_date_value(getattr(reservation, 'start_date', None)),
                'end_date': FinancialExportService._normalize_date_value(getattr(reservation, 'end_date', None)),
                'delivery_date': FinancialExportService._normalize_date_value(getattr(reservation, 'delivery_date', None)),
                'event_date': FinancialExportService._normalize_date_value(getattr(reservation, 'event_date', None)),
                'rental_duration_days': (reservation.rental_days or 0),
                'status_code': getattr(reservation, 'status', '') or 'UNKNOWN',
                'status_label': FinancialExportService._normalize_status_label(getattr(reservation, 'status', None)),
                'payment_status_code': getattr(reservation, 'payment_status', '') or 'UNKNOWN',
                'payment_status_label': FinancialExportService._normalize_payment_status_label(getattr(reservation, 'payment_status', None)),
                'rent_price': FinancialExportService._normalize_money(getattr(reservation, 'rent_price', None)),
                'discount_amount': FinancialExportService._normalize_money(getattr(reservation, 'discount_amount', None)),
                'final_price': FinancialExportService._normalize_money(getattr(reservation, 'final_price', None)),
                'deposit_amount': FinancialExportService._normalize_money(getattr(reservation, 'deposit_amount', None)),
                'remaining_amount': FinancialExportService._normalize_money(getattr(reservation, 'remaining_amount', None)),
                'refunded_amount': FinancialExportService._normalize_money(getattr(reservation, 'refunded_amount', None)),
                'cancellation_fee': FinancialExportService._normalize_money(getattr(reservation, 'cancellation_fee', None)),
                'additional_fee_total': FinancialExportService._normalize_money(additional_fee_total),
                'damage_flag': FinancialExportService._normalize_boolean_value(damage_flag),
                'damage_amount': FinancialExportService._normalize_money(getattr(reservation, 'damage_amount', None)),
                'total_collected': FinancialExportService._normalize_money(collected),
                'total_uncollected': FinancialExportService._normalize_money(remaining_balance),
                'payment_method_code': getattr(reservation, 'payment_method', '') or '-',
                'payment_tracking_code': getattr(reservation, 'payment_tracking_code', '') or '-',
                'remaining_payment_method_code': getattr(reservation, 'remaining_payment_method', '') or '-',
                'remaining_payment_tracking_code': getattr(reservation, 'remaining_payment_tracking_code', '') or '-',
                'remaining_paid_at': FinancialExportService._normalize_datetime_value(getattr(reservation, 'remaining_paid_at', None)),
                'created_by_name': FinancialExportService._user_label(getattr(reservation, 'created_by', None)),
                'updated_by_name': FinancialExportService._user_label(getattr(reservation, 'updated_by', None)),
                'archived_by_name': FinancialExportService._user_label(getattr(reservation, 'archived_by', None)),
                'requires_follow_up': FinancialExportService._normalize_boolean_value(remaining_balance > 0 or damage_flag),
            }
            rows.append(row)
        return rows

    @staticmethod
    def _build_transaction_rows(reservations):
        rows = []
        for reservation in reservations:
            for transaction in reservation.transactions.filter(transaction_status=Transaction.TransactionStatus.POSTED, is_voided=False):
                account = getattr(transaction, 'account', None)
                category = getattr(transaction, 'category', None)
                rows.append({
                    'transaction_id': transaction.pk,
                    'reservation_id': reservation.pk,
                    'transaction_type_code': getattr(transaction, 'transaction_type', '') or '-',
                    'transaction_type_label': FinancialExportService._translate_value(getattr(transaction, 'get_transaction_type_display', lambda: '')() or '-'),
                    'transaction_date': FinancialExportService._normalize_datetime_value(getattr(transaction, 'transaction_date', None)),
                    'amount': FinancialExportService._normalize_money(getattr(transaction, 'amount', None)),
                    'currency_code': getattr(transaction, 'currency', '') or 'IRR',
                    'account_code': getattr(account, 'code', '') or '-',
                    'account_name': getattr(account, 'name', '') or '-',
                    'category_code': getattr(category, 'name', '') or '-',
                    'category_label': getattr(category, 'name', '') or '-',
                    'payment_method_code': FinancialExportService._translate_value(getattr(transaction, 'payment_method', '') or '-'),
                    'payment_reference': getattr(transaction, 'payment_reference', '') or '-',
                    'transaction_status_code': getattr(transaction, 'transaction_status', '') or 'UNKNOWN',
                    'transaction_status_label': FinancialExportService._translate_value(getattr(transaction, 'transaction_status', '') or 'UNKNOWN'),
                    'created_by_name': FinancialExportService._user_label(getattr(transaction, 'created_by', None)),
                    'notes': (getattr(transaction, 'notes', '') or getattr(transaction, 'description', '') or '-'),
                    'is_voided': FinancialExportService._normalize_boolean_value(getattr(transaction, 'is_voided', False)),
                    'legacy_source': 'reservation',
                })
        return rows

    @staticmethod
    def _build_audit_rows(reservations):
        rows = []
        for reservation in reservations:
            rows.append({
                'reservation_id': reservation.pk,
                'legacy_reservation_id': reservation.pk,
                'customer_id': getattr(getattr(reservation, 'customer', None), 'pk', None),
                'dress_id': getattr(getattr(reservation, 'dress', None), 'pk', None),
                'created_by_id': getattr(getattr(reservation, 'created_by', None), 'pk', None),
                'updated_by_id': getattr(getattr(reservation, 'updated_by', None), 'pk', None),
                'archived_by_id': getattr(getattr(reservation, 'archived_by', None), 'pk', None),
                'status_code': getattr(reservation, 'status', '') or 'UNKNOWN',
                'payment_status_code': getattr(reservation, 'payment_status', '') or 'UNKNOWN',
                'payment_method_code': getattr(reservation, 'payment_method', '') or '-',
                'payment_tracking_code': getattr(reservation, 'payment_tracking_code', '') or '-',
                'remaining_payment_tracking_code': getattr(reservation, 'remaining_payment_tracking_code', '') or '-',
                'created_at': FinancialExportService._normalize_datetime_value(getattr(reservation, 'created_at', None)),
                'updated_at': FinancialExportService._normalize_datetime_value(getattr(reservation, 'updated_at', None)),
                'archived_at': FinancialExportService._normalize_datetime_value(getattr(reservation, 'archived_at', None)),
                'transaction_mapping_key': f'reservation:{reservation.pk}',
                'legacy_field_map': 'deposit_amount->DEPOSIT; remaining_payment_amount->FINAL_PAYMENT; cancellation_fee->CANCELLATION_FEE',
                'migration_notes': 'legacy reservation export',
            })
        return rows

    @staticmethod
    def _get_variant_columns(variant):
        if variant == 'accounting':
            return {
                'Summary': [
                    'export_variant', 'generated_at', 'date_from', 'date_to', 'total_reservations', 'total_final_amount',
                    'total_deposit', 'total_remaining_balance', 'total_cancellation_fees', 'cancelled_count',
                    'damaged_count', 'total_collected', 'total_uncollected', 'settlement_ratio'
                ],
                'Reservations': [
                    'reservation_id', 'contract_number', 'customer_id', 'customer_name', 'bride_phone',
                    'dress_id', 'dress_code', 'start_date', 'start_date_jalali', 'end_date', 'delivery_date', 'event_date',
                    'rental_duration_days', 'status_code', 'status_label', 'payment_status_code', 'payment_status_label',
                    'rent_price', 'discount_amount', 'final_price', 'deposit_amount', 'remaining_amount', 'refunded_amount',
                    'cancellation_fee', 'additional_fee_total', 'damage_flag', 'damage_amount', 'total_collected',
                    'total_uncollected', 'payment_method_code',
                    'payment_tracking_code', 'remaining_payment_method_code', 'remaining_payment_tracking_code',
                    'remaining_paid_at', 'created_by_name', 'updated_by_name', 'archived_by_name', 'requires_follow_up'
                ],
                'Transactions': [
                    'transaction_id', 'reservation_id', 'transaction_type_code', 'transaction_type_label', 'transaction_date',
                    'amount', 'currency_code', 'account_code', 'account_name', 'category_code', 'category_label',
                    'payment_method_code', 'payment_reference', 'transaction_status_code', 'transaction_status_label',
                    'created_by_name', 'notes', 'is_voided', 'legacy_source'
                ],
                'Audit_Migration': [
                    'reservation_id', 'transaction_id', 'legacy_reservation_id', 'customer_id', 'dress_id', 'created_by_id',
                    'updated_by_id', 'archived_by_id', 'status_code', 'payment_status_code', 'payment_method_code',
                    'payment_tracking_code', 'remaining_payment_tracking_code', 'created_at', 'updated_at', 'archived_at',
                    'transaction_mapping_key', 'legacy_field_map', 'migration_notes'
                ],
            }
        if variant == 'migration':
            return {
                'Summary': [
                    'export_variant', 'generated_at', 'date_from', 'date_to', 'total_reservations', 'total_final_amount',
                    'total_deposit', 'total_remaining_balance', 'total_cancellation_fees', 'cancelled_count',
                    'damaged_count', 'total_collected', 'total_uncollected', 'settlement_ratio'
                ],
                'Reservations': [
                    'reservation_id', 'contract_number', 'customer_id', 'customer_name', 'bride_phone',
                    'dress_id', 'dress_code', 'start_date', 'start_date_jalali', 'end_date', 'delivery_date', 'event_date',
                    'rental_duration_days', 'status_code', 'status_label', 'payment_status_code', 'payment_status_label',
                    'rent_price', 'discount_amount', 'final_price', 'deposit_amount', 'remaining_amount', 'refunded_amount',
                    'cancellation_fee', 'additional_fee_total', 'damage_flag', 'damage_amount', 'total_collected',
                    'total_uncollected', 'payment_method_code',
                    'payment_tracking_code', 'remaining_payment_method_code', 'remaining_payment_tracking_code',
                    'remaining_paid_at', 'created_by_name', 'updated_by_name', 'archived_by_name', 'requires_follow_up'
                ],
                'Transactions': [
                    'transaction_id', 'reservation_id', 'transaction_type_code', 'transaction_type_label', 'transaction_date',
                    'amount', 'currency_code', 'account_code', 'account_name', 'category_code', 'category_label',
                    'payment_method_code', 'payment_reference', 'transaction_status_code', 'transaction_status_label',
                    'created_by_name', 'notes', 'is_voided', 'legacy_source'
                ],
                'Audit_Migration': [
                    'reservation_id', 'transaction_id', 'legacy_reservation_id', 'customer_id', 'dress_id', 'created_by_id',
                    'updated_by_id', 'archived_by_id', 'status_code', 'payment_status_code', 'payment_method_code',
                    'payment_tracking_code', 'remaining_payment_tracking_code', 'created_at', 'updated_at', 'archived_at',
                    'transaction_mapping_key', 'legacy_field_map', 'migration_notes'
                ],
            }
        return {
            'Summary': [
                'export_variant', 'generated_at', 'date_from', 'date_to', 'total_reservations', 'total_final_amount',
                'total_deposit', 'total_remaining_balance', 'total_cancellation_fees', 'cancelled_count',
                'damaged_count', 'settlement_ratio'
            ],
            'Reservations': [
                'reservation_id', 'customer_name', 'bride_phone', 'dress_code', 'start_date', 'end_date',
                'status_code', 'status_label', 'payment_status_code', 'payment_status_label', 'rent_price', 'discount_amount', 'final_price',
                'deposit_amount', 'remaining_amount', 'cancellation_fee', 'additional_fee_total', 'damage_flag',
                'total_collected', 'payment_method_code', 'payment_tracking_code',
                'created_by_name'
            ],
            'Transactions': [],
            'Audit_Migration': [],
        }

    @staticmethod
    def _get_headers(sheet_name, variant):
        columns = FinancialExportService._get_variant_columns(variant)[sheet_name]
        if not columns:
            return []
        return [
            {
                'field': field,
                'label': FinancialExportService._get_header_label(field, sheet_name),
            }
            for field in columns
        ]

    @staticmethod
    def _get_header_label(field, sheet_name=None):
        if sheet_name == 'Audit_Migration':
            labels = {
                'id': 'شناسه',
                'created': 'تاریخ ایجاد',
                'updated': 'تاریخ بروزرسانی',
                'transaction_id': 'شناسه تراکنش',
                'task_id': 'شناسه وظیفه',
                'user_id': 'شناسه کاربر',
                'user_info': 'اطلاعات کاربر',
                'reservation_id': 'شناسه',
                'legacy_reservation_id': 'شناسه رزرو قدیمی',
                'customer_id': 'شناسه مشتری',
                'dress_id': 'شناسه لباس',
                'created_by_id': 'شناسه ایجاد کننده',
                'updated_by_id': 'شناسه ویرایش کننده',
                'archived_by_id': 'شناسه آرشیو کننده',
                'status_code': 'وضعیت',
                'payment_status_code': 'وضعیت پرداخت',
                'payment_method_code': 'روش پرداخت',
                'payment_tracking_code': 'کد رهگیری پرداخت',
                'remaining_payment_tracking_code': 'کد رهگیری پرداخت باقی‌مانده',
                'created_at': 'تاریخ ایجاد',
                'updated_at': 'تاریخ بروزرسانی',
                'archived_at': 'تاریخ آرشیو',
                'transaction_mapping_key': 'کلید نگاشت تراکنش',
                'legacy_field_map': 'نقشه فیلدهای قدیمی',
                'migration_notes': 'یادداشت مهاجرت',
            }
        elif sheet_name == 'Transactions':
            labels = {
                'transaction_id': 'شناسه تراکنش',
                'reservation_id': 'شناسه رزرو',
                'transaction_type_code': 'نوع تراکنش',
                'transaction_type_label': 'شرح نوع تراکنش',
                'transaction_date': 'تاریخ تراکنش',
                'amount': 'مبلغ',
                'currency_code': 'ارز',
                'account_code': 'کد حساب',
                'account_name': 'نام حساب',
                'category_code': 'کد دسته',
                'category_label': 'دسته',
                'payment_method_code': 'روش پرداخت',
                'payment_reference': 'مرجع پرداخت',
                'transaction_status_code': 'وضعیت تراکنش',
                'transaction_status_label': 'شرح وضعیت تراکنش',
                'created_by_name': 'ثبت کننده',
                'notes': 'یادداشت',
                'is_voided': 'ابطال شده؟',
                'legacy_source': 'منبع قدیمی',
            }
        else:
            labels = {
                'reservation_id': 'شناسه رزرو',
                'contract_number': 'شماره قرارداد',
                'customer_id': 'شناسه مشتری',
                'customer_name': 'نام مشتری',
                'bride_phone': 'تلفن عروس',
                'dress_id': 'شناسه لباس',
                'dress_code': 'کد لباس',
                'start_date': 'تاریخ شروع',
                'start_date_jalali': 'تاریخ شروع',
                'end_date': 'تاریخ پایان',
                'delivery_date': 'تاریخ تحویل',
                'event_date': 'تاریخ مراسم',
                'rental_duration_days': 'روزهای اجاره',
                'status_code': 'وضعیت',
                'status_label': 'وضعیت',
                'payment_status_code': 'وضعیت پرداخت',
                'payment_status_label': 'وضعیت پرداخت',
                'rent_price': 'هزینه اجاره',
                'discount_amount': 'تخفیف',
                'final_price': 'مبلغ نهایی',
                'deposit_amount': 'بیعانه',
                'remaining_amount': 'باقی‌مانده',
                'refunded_amount': 'مبلغ مرجوعی',
                'cancellation_fee': 'هزینه لغو',
                'additional_fee_total': 'جمع هزینه‌های جانبی',
                'damage_flag': 'آسیب؟',
                'damage_amount': 'مبلغ آسیب',
                'total_collected': 'جمع دریافتی',
                'total_uncollected': 'جمع مانده',
                'payment_method_code': 'روش پرداخت',
                'payment_tracking_code': 'کد رهگیری پرداخت',
                'remaining_payment_method_code': 'روش پرداخت باقی‌مانده',
                'remaining_payment_tracking_code': 'کد رهگیری پرداخت باقی‌مانده',
                'remaining_paid_at': 'تاریخ پرداخت باقی‌مانده',
                'created_by_name': 'ثبت کننده',
                'updated_by_name': 'ویرایش کننده',
                'archived_by_name': 'آرشیو کننده',
                'requires_follow_up': 'نیاز به پیگیری',
                'created_at': 'تاریخ ایجاد',
                'updated_at': 'تاریخ بروزرسانی',
                'archived_at': 'تاریخ آرشیو',
                'legacy_reservation_id': 'شناسه رزرو قدیمی',
                'transaction_mapping_key': 'کلید نگاشت تراکنش',
                'legacy_field_map': 'نقشه فیلدهای قدیمی',
                'migration_notes': 'یادداشت مهاجرت',
            }
        return labels.get(field, field.replace('_', ' ').replace('id', 'ID').title())

    @staticmethod
    def _build_clean_headers(columns, sheet_name=None):
        headers = []
        seen = set()
        for column in columns:
            label = FinancialExportService._get_header_label(column, sheet_name)
            if not label:
                continue
            if label in seen:
                continue
            seen.add(label)
            headers.append((column, label))
        return headers

    @staticmethod
    def _sheet_title(sheet_name):
        return FinancialExportService.SHEET_DEFINITIONS.get(sheet_name, {}).get('title', sheet_name)

    @staticmethod
    def _get_summary_label(metric):
        labels = {
            'export_variant': 'نوع خروجی',
            'generated_at': 'زمان ساخت فایل',
            'date_from': 'از تاریخ',
            'date_to': 'تا تاریخ',
            'total_reservations': 'تعداد رزروها',
            'total_final_amount': 'جمع مبلغ نهایی',
            'total_deposit': 'جمع بیعانه',
            'total_remaining_balance': 'جمع مانده',
            'total_cancellation_fees': 'جمع جریمه لغو',
            'cancelled_count': 'تعداد لغو شده',
            'damaged_count': 'تعداد آسیب‌دیده',
            'total_collected': 'جمع دریافتی',
            'total_uncollected': 'جمع معوق',
            'settlement_ratio': 'نرخ تسویه',
        }
        return labels.get(metric, metric.replace('_', ' ').title())

    @staticmethod
    def _format_summary_value(metric, value):
        if metric in {'total_final_amount', 'total_deposit', 'total_remaining_balance', 'total_cancellation_fees', 'total_collected', 'total_uncollected'}:
            if value in (None, ''):
                return 0
            try:
                return int(value)
            except (TypeError, ValueError):
                return value
        if metric == 'settlement_ratio':
            if value in (None, ''):
                return 0
            try:
                return round(float(value) * 100, 2)
            except (TypeError, ValueError):
                return value
        return value

    @staticmethod
    def _apply_sheet_style(ws):
        header_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
        header_font = Font(bold=True)
        title_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        title_font = Font(bold=True, color='FFFFFF', size=14)
        section_fill = PatternFill(fill_type='solid', fgColor='EAF2F8')
        thin = Side(border_style='thin', color='D0D7DE')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        alignment = Alignment(horizontal='center', vertical='center')

        if ws.title == FinancialExportService.SHEET_DEFINITIONS['Summary']['title'] and ws.max_row >= 3:
            ws['A1'].value = 'خلاصه مدیریتی'
            ws['A1'].fill = title_fill
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A1'].border = border
            ws['B1'].fill = title_fill
            ws['B1'].font = title_font
            ws['B1'].border = border
            ws.merge_cells('A1:B1')
            ws['A2'].fill = section_fill
            ws['A2'].font = Font(bold=True)
            ws['A2'].border = border
            ws['B2'].fill = section_fill
            ws['B2'].border = border
            ws['A3'].fill = header_fill
            ws['A3'].font = header_font
            ws['A3'].alignment = alignment
            ws['A3'].border = border
            ws['B3'].fill = header_fill
            ws['B3'].font = header_font
            ws['B3'].alignment = alignment
            ws['B3'].border = border

        header_row = 1 if ws.title != FinancialExportService.SHEET_DEFINITIONS['Summary']['title'] else 3
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border

        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for row_index in range(1, ws.max_row + 1):
            ws.row_dimensions[row_index].height = 20

        for column_index in range(1, ws.max_column + 1):
            column_letter = get_column_letter(column_index)
            values = [cell.value for cell in ws[column_letter] if cell.value is not None]
            if not values:
                continue
            max_length = max(len(str(value)) for value in values)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        ws.freeze_panes = 'A2' if ws.title != FinancialExportService.SHEET_DEFINITIONS['Summary']['title'] else 'A4'
        ws.sheet_view.rightToLeft = True

    @staticmethod
    def build_workbook(filters=None, variant='management'):
        variant = FinancialExportService.normalize_variant(variant)
        filters = filters or {}
        queryset = FinancialExportService._build_queryset(filters)
        reservations = list(queryset)

        wb = Workbook()
        for sheet_name in ['Reservations', 'Summary', 'Transactions', 'Audit_Migration']:
            ws = wb.create_sheet(title=FinancialExportService._sheet_title(sheet_name))
            if sheet_name == 'Summary':
                summary_rows = FinancialExportService._build_summary_rows(reservations, variant, filters=filters)
                ws.append(['خلاصه مدیریتی', ''])
                ws.append(['بازه زمانی', f"{filters.get('date_from') or '-'} تا {filters.get('date_to') or '-'}"])
                ws.append(['متریک', 'ارزش'])
                for key, value in summary_rows:
                    ws.append([FinancialExportService._get_summary_label(key), FinancialExportService._format_summary_value(key, value)])
            elif sheet_name == 'Reservations':
                columns = FinancialExportService._get_variant_columns(variant)[sheet_name]
                headers = FinancialExportService._build_clean_headers(columns, sheet_name)
                ws.append([label for _, label in headers])
                for row in FinancialExportService._build_reservation_rows(reservations, variant):
                    values = [FinancialExportService._resolve_row_value(row, field) for field, _ in headers]
                    ws.append(values)
            elif sheet_name == 'Transactions':
                if variant in {'accounting', 'migration'}:
                    columns = FinancialExportService._get_variant_columns(variant)[sheet_name]
                    headers = FinancialExportService._build_clean_headers(columns, sheet_name)
                    ws.append([label for _, label in headers])
                    for row in FinancialExportService._build_transaction_rows(reservations):
                        values = [FinancialExportService._resolve_row_value(row, field) for field, _ in headers]
                        ws.append(values)
            elif sheet_name == 'Audit_Migration':
                if variant in {'accounting', 'migration'}:
                    columns = FinancialExportService._get_variant_columns(variant)[sheet_name]
                    headers = FinancialExportService._build_clean_headers(columns, sheet_name)
                    ws.append([label for _, label in headers])
                    for row in FinancialExportService._build_audit_rows(reservations):
                        values = [FinancialExportService._resolve_row_value(row, field) for field, _ in headers]
                        ws.append(values)
            FinancialExportService._apply_sheet_style(ws)

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
