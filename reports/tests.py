from io import BytesIO

import jdatetime
from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from reservations.constants import ReservationStatus


class ReportsIndexViewTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='report-admin',
            password='test-pass123',
            role=get_user_model().Role.MANAGER,
        )

    def test_reports_index_renders_operational_summary(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('reports:index'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'خلاصه عملکرد و گزارش‌های مدیریتی')
        self.assertContains(response, 'فیلتر گزارش')
        self.assertContains(response, 'روند درآمد و رزرو')

    def test_reports_filters_accept_dress_filter_without_error(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('reports:index'), {'dress_id': 999})

        self.assertEqual(response.status_code, 200)

    def test_reports_filter_inputs_use_persian_datepicker(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('reports:index'))

        self.assertContains(response, 'id="dateFromInput"')
        self.assertContains(response, 'id="dateToInput"')
        self.assertContains(response, 'class="form-control form-control-sm p-date-only"')

    def test_reports_summary_uses_jalali_labels_for_date_range(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('reports:index'), {'date_from': '2026-07-08', 'date_to': '2026-07-15'})

        self.assertEqual(response.status_code, 200)
        self.assertIn('1405/04/17 تا 1405/04/24', response.context['summary']['period_label'])

    def test_reports_date_presets_include_future_week_and_month(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('reports:index'))

        self.assertEqual(response.status_code, 200)
        labels = [preset['label'] for preset in response.context['date_presets']]
        self.assertIn('هفته آینده', labels)
        self.assertIn('ماه آینده', labels)

    def test_reports_export_without_filters_returns_all_reservations(self):
        from customers.models import Customer
        from products.models import Dress
        from reservations.models import Reservation

        customer = Customer.objects.create(
            bride_first_name='علیه',
            bride_last_name='مهدی',
            bride_phone='09120000000',
            ceremony_date=jdatetime.date(1402, 1, 1),
            how_to_know='test',
            allow_contact=True,
        )
        dress = Dress.objects.create(code='REPORT1', daily_rent_price=100000)
        Reservation.objects.create(
            customer=customer,
            dress=dress,
            created_by=self.user,
            start_date=jdatetime.date(1405, 4, 10),
            rental_days=4,
            rent_price=100000,
            deposit_amount=50000,
            discount_type='NONE',
            discount_value=0,
            discount_amount=0,
            final_price=100000,
            remaining_payment_amount=50000,
            remaining_amount=50000,
            refunded_amount=0,
            payment_status=Reservation.PAYMENT_PARTIAL,
            payment_method='CASH',
            payment_tracking_code='TK-1',
            guarantee1_type='CASH',
            guarantee1_tracking_code='G1',
        )
        Reservation.objects.create(
            customer=customer,
            dress=dress,
            created_by=self.user,
            start_date=jdatetime.date(1405, 3, 11),
            rental_days=2,
            rent_price=150000,
            deposit_amount=100000,
            discount_type='NONE',
            discount_value=0,
            discount_amount=0,
            final_price=150000,
            remaining_payment_amount=50000,
            remaining_amount=50000,
            refunded_amount=0,
            payment_status=Reservation.PAYMENT_PARTIAL,
            payment_method='CASH',
            payment_tracking_code='TK-2',
            guarantee1_type='CASH',
            guarantee1_tracking_code='G2',
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse('reports:export_excel'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
        worksheet = workbook['جزئیات رزروها']
        rows = list(worksheet.iter_rows(values_only=True))

        self.assertGreaterEqual(len(rows), 5)
        self.assertEqual(rows[0][0], 'جزئیات رزروها')
        self.assertEqual(rows[4][0], 'شناسه رزرو')

    def test_reports_export_applies_professional_excel_styling(self):
        from customers.models import Customer
        from products.models import Dress
        from reservations.models import Reservation

        customer = Customer.objects.create(
            bride_first_name='سارا',
            bride_last_name='رضایی',
            bride_phone='09120000000',
            ceremony_date=jdatetime.date(1402, 1, 1),
            how_to_know='test',
            allow_contact=True,
        )
        dress = Dress.objects.create(code='REPORT3', daily_rent_price=130000)
        Reservation.objects.create(
            customer=customer,
            dress=dress,
            created_by=self.user,
            start_date=jdatetime.date(1405, 4, 10),
            rental_days=4,
            rent_price=130000,
            deposit_amount=65000,
            discount_type='NONE',
            discount_value=0,
            discount_amount=0,
            final_price=130000,
            remaining_payment_amount=65000,
            remaining_amount=65000,
            refunded_amount=0,
            payment_status=Reservation.PAYMENT_PARTIAL,
            payment_method='CASH',
            payment_tracking_code='TK-4',
            status=ReservationStatus.CONFIRMED,
            guarantee1_type='CASH',
            guarantee1_tracking_code='G4',
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse('reports:export_excel'), {'report_type': 'details'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
        worksheet = workbook['جزئیات رزروها']

        self.assertEqual(worksheet.freeze_panes, 'A6')
        self.assertEqual(worksheet.auto_filter.ref, f'A5:{get_column_letter(worksheet.max_column)}{worksheet.max_row}')
        self.assertEqual(worksheet['A1'].value, 'جزئیات رزروها')
        self.assertTrue(worksheet['A1'].font.bold)
        self.assertTrue(worksheet['A5'].font.bold)

    def test_reports_export_uses_persian_labels_and_management_sheets(self):
        from customers.models import Customer
        from products.models import Dress
        from reservations.models import Reservation

        customer = Customer.objects.create(
            bride_first_name='سارا',
            bride_last_name='رضایی',
            bride_phone='09120000000',
            ceremony_date=jdatetime.date(1402, 1, 1),
            how_to_know='test',
            allow_contact=True,
        )
        dress = Dress.objects.create(code='REPORT2', daily_rent_price=120000)
        Reservation.objects.create(
            customer=customer,
            dress=dress,
            created_by=self.user,
            start_date=jdatetime.date(1405, 4, 10),
            rental_days=4,
            rent_price=120000,
            deposit_amount=60000,
            discount_type='NONE',
            discount_value=0,
            discount_amount=0,
            final_price=120000,
            remaining_payment_amount=60000,
            remaining_amount=60000,
            refunded_amount=0,
            payment_status=Reservation.PAYMENT_PARTIAL,
            payment_method='CASH',
            payment_tracking_code='TK-3',
            status=ReservationStatus.CONFIRMED,
            guarantee1_type='CASH',
            guarantee1_tracking_code='G3',
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse('reports:export_excel'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
        self.assertIn('خلاصه مدیریتی', workbook.sheetnames)
        self.assertIn('جزئیات رزروها', workbook.sheetnames)

        summary_sheet = workbook['خلاصه مدیریتی']
        values = [cell.value for row in summary_sheet.iter_rows(min_row=1, max_row=20) for cell in row]
        self.assertIn('متریک', values)
        self.assertIn('رزروهای فعال', values)

        reservation_sheet = workbook['جزئیات رزروها']
        headers = [cell.value for cell in next(reservation_sheet.iter_rows(min_row=5, max_row=5))]
        self.assertIn('وضعیت رزرو', headers)
        self.assertIn('روش پرداخت', headers)

        first_row = next(reservation_sheet.iter_rows(min_row=6, max_row=6))
        values = [cell.value for cell in first_row]
        self.assertIn('قطعی', values)
        self.assertIn('پرداخت جزئی', values)
        self.assertIn('نقدی', values)
