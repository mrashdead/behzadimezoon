from datetime import date, datetime, timedelta
from io import BytesIO
from urllib.parse import urlencode
import jdatetime

from django.contrib.auth.decorators import login_required
from django.db.models import Case, Count, IntegerField, Q, Sum, When
from django.http import HttpResponse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from accounts.models import User
from customers.models import Customer
from financial.models import CancellationRecord, DamageRecord, Transaction
from financial.services.dashboard_service import DashboardService
from products.models import Dress
from reservations.constants import PaymentMethod, ReservationStatus
from reservations.models import AdditionalFee, Reservation
from reservations.utils import parse_reservation_date


EXPORT_VALUE_TRANSLATIONS = {
    'CONFIRMED': 'قطعی',
    'CANCELLED': 'لغو شده',
    'DELIVERED': 'تحویل شده',
    'PARTIAL': 'پرداخت جزئی',
    'PAID': 'پرداخت کامل',
    'UNPAID': 'پرداخت نشده',
    'REFUNDED': 'بازپرداخت',
    'No': 'خیر',
    'Yes': 'بله',
    'CASH': 'نقدی',
    'TRANSFER': 'انتقال بانکی',
    'POS': 'کارتخوان',
    'CARD': 'کارت به کارت',
    'PAYMENT': 'پرداخت',
    'DEPOSIT': 'بیعانه',
    'REFUND': 'بازپرداخت',
    'CANCELLATION_FEE': 'جریمه لغو',
    'DAMAGE_CHARGE': 'هزینه خسارت',
    'ACCRUAL': 'تعهدی',
    'UNKNOWN': 'ناشناخته',
}


def _persian_digits(value):
    if value in (None, ''):
        return ''
    text = str(value)
    return text.translate(str.maketrans('0123456789', '۰۱۲۳۴۵۶۷۸۹'))


def _clean_text(value):
    if value in (None, ''):
        return ''
    if isinstance(value, bool):
        return 'بله' if value else 'خیر'
    text = str(value).strip()
    return ' '.join(text.split())


def _translate_export_value(value):
    if value in (None, ''):
        return ''
    if isinstance(value, bool):
        return 'بله' if value else 'خیر'
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return ''
    translated = EXPORT_VALUE_TRANSLATIONS.get(text, EXPORT_VALUE_TRANSLATIONS.get(text.upper(), None))
    if translated is not None:
        return translated

    try:
        from reservations.constants import ReservationStatus, PaymentMethod
        status_lookup = dict(ReservationStatus.CHOICES)
        method_lookup = dict(PaymentMethod.CHOICES)
        if text in status_lookup:
            return status_lookup[text]
        if text in method_lookup:
            return method_lookup[text]
    except Exception:
        pass

    return _clean_text(text)


def _format_export_date(value):
    if value in (None, ''):
        return ''
    if isinstance(value, datetime):
        try:
            return _persian_digits(jdatetime.datetime.fromgregorian(datetime=value).strftime('%Y/%m/%d'))
        except Exception:
            return _persian_digits(value.strftime('%Y/%m/%d'))
    if isinstance(value, date):
        try:
            return _persian_digits(jdatetime.date.fromgregorian(date=value).strftime('%Y/%m/%d'))
        except Exception:
            return _persian_digits(value.strftime('%Y/%m/%d'))
    if isinstance(value, str):
        parsed = parse_reservation_date(value)
        if parsed:
            return _format_export_date(parsed)
        return _clean_text(value)
    if hasattr(value, 'togregorian'):
        try:
            return _persian_digits(value.strftime('%Y/%m/%d'))
        except Exception:
            return _clean_text(value)
    return _clean_text(value)


def _format_export_datetime(value):
    if value in (None, ''):
        return ''
    if isinstance(value, datetime):
        try:
            return _persian_digits(jdatetime.datetime.fromgregorian(datetime=value).strftime('%Y/%m/%d %H:%M:%S'))
        except Exception:
            return _persian_digits(value.strftime('%Y/%m/%d %H:%M:%S'))
    if isinstance(value, date):
        return _format_export_date(value)
    return _clean_text(value)


def _sanitize_tracking_code(value):
    if value in (None, ''):
        return '-'
    text = str(value).strip().replace('\n', ' ').replace('\r', ' ')
    text = ''.join(ch for ch in text if ch not in {'\t', '\u200c'})
    return ' '.join(text.split()) or '-'


def _user_label(user):
    if not user:
        return '-'
    if hasattr(user, 'get_full_name'):
        name = user.get_full_name() or ''
        if name:
            return _clean_text(name)
    if hasattr(user, 'username') and user.username:
        return _clean_text(user.username)
    return _clean_text(str(user))


def _excel_display_value(value):
    if value in (None, ''):
        return '-'
    if isinstance(value, bool):
        return 'بله' if value else 'خیر'
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value).strip()
    return text if text else '-'


def _excel_number_format_for_header(header):
    label = str(header or '').strip()
    if any(keyword in label for keyword in ['مبلغ', 'درآمد', 'قیمت', 'بیعانه', 'مانده', 'خسارت', 'جریمه', 'بازپرداخت', 'نهایی', 'هزینه', 'جمع']):
        return '#,##0'
    if any(keyword in label for keyword in ['درصد', 'نرخ']):
        return '0.0%'
    return '0'


def _excel_column_alignment(header):
    label = str(header or '').strip()
    if any(keyword in label for keyword in ['مبلغ', 'درآمد', 'قیمت', 'بیعانه', 'مانده', 'خسارت', 'جریمه', 'بازپرداخت', 'نهایی', 'هزینه', 'جمع', 'تعداد', 'روز', 'رتبه']):
        return 'right'
    return 'left'


def _append_excel_rows(ws, title_text, headers, rows, metadata=None):
    ws.append([title_text])
    for label, value in (metadata or []):
        ws.append([label, value])
    ws.append([])
    ws.append([_excel_display_value(header) for header in headers])
    for row in rows:
        ws.append([_excel_display_value(value) for value in row])


def _apply_excel_sheet_style(ws, title_text, header_row, metadata_count=0):
    ws.sheet_view.rightToLeft = True

    title_fill = PatternFill(fill_type='solid', fgColor='E8F1FB')
    title_font = Font(bold=True, size=14, color='1F4E78')
    header_fill = PatternFill(fill_type='solid', fgColor='DDEBF7')
    header_font = Font(bold=True, color='1F2937')
    meta_fill = PatternFill(fill_type='solid', fgColor='F7FAFC')
    alt_fill = PatternFill(fill_type='solid', fgColor='FCFDFF')
    thin = Side(border_style='thin', color='D0D7DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_cell = ws['A1']
    title_cell.value = title_text
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_cell.border = border

    for row_index in range(2, header_row):
        for cell in ws[row_index]:
            cell.fill = meta_fill
            cell.font = Font(bold=False, color='334155')
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = border

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for row_index in range(header_row + 1, ws.max_row + 1):
        row_fill = alt_fill if row_index % 2 == 0 else None
        for col_index, cell in enumerate(ws[row_index], start=1):
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if row_fill:
                cell.fill = row_fill
            if cell.value in (None, '-', ''):
                cell.value = '-'
            header_value = ws.cell(row=header_row, column=col_index).value
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                if _excel_column_alignment(header_value) == 'right':
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.number_format = _excel_number_format_for_header(header_value)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for column_index in range(1, ws.max_column + 1):
        column_letter = get_column_letter(column_index)
        values = [cell.value for cell in ws[column_letter] if cell.value is not None]
        if not values:
            continue
        max_length = max(len(str(value)) for value in values)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}'
    ws.freeze_panes = f'A{header_row + 1}'


@method_decorator(login_required, name='dispatch')
class ReportsIndexView(TemplateView):
    template_name = 'reports/index.html'

    def _parse_date(self, value):
        if not value:
            return None
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()
        parsed = parse_reservation_date(value)
        if parsed:
            return parsed
        try:
            return date.fromisoformat(str(value))
        except ValueError:
            return None

    def _gregorian_to_jalali(self, gregorian_date):
        """Convert Gregorian date to Jalali date string (YYYY/MM/DD format)"""
        if not gregorian_date:
            return ''
        try:
            j_date = jdatetime.date.fromgregorian(date=gregorian_date)
            return j_date.strftime('%Y/%m/%d')
        except Exception:
            return ''

    def _format_jalali_date(self, value):
        """Format a date-like value as Jalali string for the report UI."""
        if not value:
            return ''
        if isinstance(value, datetime):
            value = value.date()
        if isinstance(value, date):
            try:
                return jdatetime.date.fromgregorian(date=value).strftime('%Y/%m/%d')
            except Exception:
                return ''
        if isinstance(value, str):
            parsed = self._parse_date(value)
            if parsed:
                return self._format_jalali_date(parsed)
            return value.strip()
        if hasattr(value, 'togregorian'):
            return value.strftime('%Y/%m/%d')
        return ''

    def _coerce_datetime_filter(self, value, *, end_of_day=False):
        """Return an aware datetime for DateTimeField filtering."""
        if not value:
            return None
        if isinstance(value, datetime):
            if timezone.is_aware(value):
                return value
            return timezone.make_aware(value, timezone.get_current_timezone())
        if isinstance(value, date):
            naive_dt = datetime.combine(value, datetime.max.time() if end_of_day else datetime.min.time())
            return timezone.make_aware(naive_dt, timezone.get_current_timezone())
        parsed = self._parse_date(value)
        if parsed:
            naive_dt = datetime.combine(parsed, datetime.max.time() if end_of_day else datetime.min.time())
            return timezone.make_aware(naive_dt, timezone.get_current_timezone())
        return None

    def _build_filter_url(self, override_params=None):
        params = {key: value for key, value in self.request.GET.items() if value}
        if override_params:
            params.update({key: value for key, value in override_params.items() if value not in (None, '')})
        if not params:
            return self.request.path
        return f"{self.request.path}?{urlencode(params)}"

    def _apply_filters(self, queryset, filters, *, date_field='start_date', field_overrides=None):
        if not queryset:
            return queryset
        field_overrides = field_overrides or {}
        if filters.get('seller_id'):
            seller_field = field_overrides.get('seller_id', 'created_by_id')
            queryset = queryset.filter(**{seller_field: filters['seller_id']})
        if filters.get('customer_id'):
            customer_field = field_overrides.get('customer_id', 'customer_id')
            queryset = queryset.filter(**{customer_field: filters['customer_id']})
        if filters.get('dress_id'):
            dress_field = field_overrides.get('dress_id', 'dress_id')
            queryset = queryset.filter(**{dress_field: filters['dress_id']})
        if filters.get('status'):
            status_field = field_overrides.get('status', 'status')
            queryset = queryset.filter(**{status_field: filters['status']})
        if filters.get('payment_method'):
            payment_field = field_overrides.get('payment_method', 'payment_method')
            queryset = queryset.filter(**{payment_field: filters['payment_method']})
        date_from = filters.get('date_from')
        date_to = filters.get('date_to')
        if date_from:
            queryset = queryset.filter(**{f'{date_field}__gte': date_from})
        if date_to:
            queryset = queryset.filter(**{f'{date_field}__lte': date_to})
        return queryset

    def _active_filter_labels(self, filters, selected_seller, selected_customer, selected_dress):
        labels = []
        if filters.get('date_from'):
            labels.append(f"از {self._format_jalali_date(filters['date_from'])}")
        if filters.get('date_to'):
            labels.append(f"تا {self._format_jalali_date(filters['date_to'])}")
        if selected_seller:
            labels.append(f"فروشنده: {selected_seller}")
        if selected_customer:
            labels.append(f"مشتری: {selected_customer}")
        if selected_dress:
            labels.append(f"لباس: {selected_dress}")
        if filters.get('status'):
            labels.append(f"وضعیت: {dict(ReservationStatus.CHOICES).get(filters['status'], filters['status'])}")
        if filters.get('payment_method'):
            labels.append(f"پرداخت: {dict(PaymentMethod.CHOICES).get(filters['payment_method'], filters['payment_method'])}")
        return labels

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request
        today = timezone.localdate()
        today_jalali = jdatetime.date.fromgregorian(date=today)
        current_month_start = jdatetime.date(today_jalali.year, today_jalali.month, 1).togregorian()

        date_from = self._parse_date(request.GET.get('date_from'))
        date_to = self._parse_date(request.GET.get('date_to'))
        if not date_from:
            date_from = current_month_start
        if not date_to:
            date_to = today

        filters = {
            'date_from': date_from,
            'date_to': date_to,
            'seller_id': request.GET.get('seller_id') or '',
            'customer_id': request.GET.get('customer_id') or '',
            'dress_id': request.GET.get('dress_id') or '',
            'status': request.GET.get('status') or '',
            'payment_method': request.GET.get('payment_method') or '',
        }
        transaction_filters = dict(filters)
        transaction_filters['date_from'] = self._coerce_datetime_filter(filters['date_from'])
        transaction_filters['date_to'] = self._coerce_datetime_filter(filters['date_to'], end_of_day=True)

        reservation_qs = self._apply_filters(Reservation.objects.filter(is_deleted=False), filters)
        transaction_qs = self._apply_filters(
            Transaction.objects.filter(transaction_status=Transaction.TransactionStatus.POSTED, is_voided=False),
            transaction_filters,
            date_field='transaction_date',
            field_overrides={
                'seller_id': 'reservation__created_by_id',
                'customer_id': 'reservation__customer_id',
                'dress_id': 'reservation__dress_id',
                'status': 'reservation__status',
                'payment_method': 'reservation__payment_method',
            },
        )

        active_statuses = [
            ReservationStatus.CONFIRMED,
            ReservationStatus.DELIVERED,
            ReservationStatus.RETURNED,
            ReservationStatus.LAUNDRY,
            ReservationStatus.READY,
        ]

        month_revenue = reservation_qs.aggregate(total=Sum('final_price'))['total'] or 0
        active_reservations = reservation_qs.filter(status__in=active_statuses).count()
        cancelled_reservations = reservation_qs.filter(status=ReservationStatus.CANCELLED).count()
        refund_cases = transaction_qs.filter(transaction_type=Transaction.TransactionType.REFUND).count()
        refund_amount = transaction_qs.filter(transaction_type=Transaction.TransactionType.REFUND).aggregate(total=Sum('amount'))['total'] or 0
        damage_penalty_total = (
            reservation_qs.aggregate(total_damage=Sum('damage_amount'))['total_damage'] or 0
        ) + (
            reservation_qs.aggregate(total_penalty=Sum('cancellation_fee'))['total_penalty'] or 0
        )

        top_dress = reservation_qs.values('dress__code').annotate(
            total_revenue=Sum('final_price'),
            reservation_count=Count('id'),
        ).order_by('-total_revenue').first()
        top_customer = reservation_qs.values(
            'customer__pk', 'customer__bride_first_name', 'customer__bride_last_name'
        ).annotate(
            total_revenue=Sum('final_price'),
            reservation_count=Count('id'),
        ).order_by('-total_revenue').first()

        cash_inflow = transaction_qs.filter(
            transaction_type__in=[
                Transaction.TransactionType.DEPOSIT,
                Transaction.TransactionType.FINAL_PAYMENT,
                Transaction.TransactionType.PARTIAL_PAYMENT,
                Transaction.TransactionType.DAMAGE_PAYMENT,
                Transaction.TransactionType.PENALTY_INCOME,
                Transaction.TransactionType.TRANSFER_IN,
                Transaction.TransactionType.ADJUSTMENT_IN,
            ]
        ).aggregate(total=Sum('amount'))['total'] or 0
        cancellation_fee_income = transaction_qs.filter(transaction_type=Transaction.TransactionType.CANCELLATION_FEE).aggregate(total=Sum('amount'))['total'] or 0
        damage_income = (
            transaction_qs.filter(transaction_type=Transaction.TransactionType.DAMAGE_PAYMENT).aggregate(total=Sum('amount'))['total'] or 0
        ) + (
            transaction_qs.filter(transaction_type=Transaction.TransactionType.PENALTY_INCOME).aggregate(total=Sum('amount'))['total'] or 0
        )
        refund_total = transaction_qs.filter(transaction_type=Transaction.TransactionType.REFUND).aggregate(total=Sum('amount'))['total'] or 0
        outstanding_total = reservation_qs.aggregate(total=Sum('remaining_amount'))['total'] or 0
        additional_fee_total = AdditionalFee.objects.filter(is_deleted=False, reservation__in=reservation_qs).aggregate(total=Sum('amount'))['total'] or 0
        adjustment_total = transaction_qs.filter(
            transaction_type__in=[Transaction.TransactionType.DISCOUNT, Transaction.TransactionType.ADJUSTMENT_IN, Transaction.TransactionType.ADJUSTMENT_OUT]
        ).aggregate(total=Sum('amount'))['total'] or 0

        repeat_customers = reservation_qs.values(
            'customer__pk', 'customer__bride_first_name', 'customer__bride_last_name'
        ).annotate(
            reservation_count=Count('id'),
            total_revenue=Sum('final_price'),
        ).order_by('-reservation_count', '-total_revenue')[:8]

        top_dresses = reservation_qs.values('dress__pk', 'dress__code').annotate(
            reservation_count=Count('id'),
            total_revenue=Sum('final_price'),
        ).order_by('-total_revenue')[:8]

        seller_rows = reservation_qs.values('created_by__id', 'created_by__username').annotate(
            reservation_count=Count('id'),
            total_revenue=Sum('final_price'),
            cancelled_count=Sum(Case(When(status=ReservationStatus.CANCELLED, then=1), default=0, output_field=IntegerField())),
            damage_total=Sum('damage_amount'),
        ).order_by('-total_revenue', '-reservation_count')[:8]

        damage_records = DamageRecord.objects.filter(reservation__in=reservation_qs).select_related('customer', 'dress').order_by('-amount')[:6]
        cancellation_records = CancellationRecord.objects.filter(reservation__in=reservation_qs).select_related('reservation').order_by('-penalty_amount')[:6]
        cancellation_total = cancellation_records.count()
        damage_total_amount = damage_records.aggregate(total=Sum('amount'))['total'] or 0

        trend_days = 14
        if date_from and date_to:
            span = (date_to - date_from).days + 1
            trend_days = max(7, min(30, span))

        labels = []
        revenue_series = []
        reservation_series = []
        current_day = date_to
        for offset in range(trend_days):
            day = current_day - timedelta(days=trend_days - 1 - offset)
            labels.append(self._format_jalali_date(day))
            day_reservations = reservation_qs.filter(start_date=day)
            revenue_series.append(day_reservations.aggregate(total=Sum('final_price'))['total'] or 0)
            reservation_series.append(day_reservations.count())

        previous_start = date_from - timedelta(days=max(1, trend_days))
        previous_end = date_from - timedelta(days=1)
        previous_reservations = Reservation.objects.filter(
            is_deleted=False,
            start_date__gte=previous_start,
            start_date__lte=previous_end,
        )
        if filters.get('seller_id'):
            previous_reservations = previous_reservations.filter(created_by_id=filters['seller_id'])
        if filters.get('customer_id'):
            previous_reservations = previous_reservations.filter(customer_id=filters['customer_id'])
        if filters.get('dress_id'):
            previous_reservations = previous_reservations.filter(dress_id=filters['dress_id'])
        if filters.get('status'):
            previous_reservations = previous_reservations.filter(status=filters['status'])
        if filters.get('payment_method'):
            previous_reservations = previous_reservations.filter(payment_method=filters['payment_method'])
        previous_cancelled = previous_reservations.filter(status=ReservationStatus.CANCELLED).count()
        current_cancelled = cancelled_reservations
        if current_cancelled > previous_cancelled:
            cancellation_trend = 'در حال افزایش است'
        else:
            cancellation_trend = 'پایدار یا کاهشی'

        previous_damage = previous_reservations.aggregate(total=Sum('damage_amount'))['total'] or 0
        current_damage = damage_penalty_total
        if current_damage > previous_damage:
            damage_trend = 'در حال افزایش است'
        else:
            damage_trend = 'پایدار یا کاهشی'

        selected_seller = User.objects.filter(pk=filters['seller_id']).first() if filters.get('seller_id') else None
        selected_customer = Customer.objects.filter(pk=filters['customer_id']).first() if filters.get('customer_id') else None
        selected_dress = Dress.objects.filter(pk=filters['dress_id']).first() if filters.get('dress_id') else None

        context['page_title'] = 'گزارش‌های عملیاتی'
        context['filters'] = filters
        context['date_from_value'] = self._format_jalali_date(request.GET.get('date_from') or date_from)
        context['date_to_value'] = self._format_jalali_date(request.GET.get('date_to') or date_to)
        context['active_filter_labels'] = self._active_filter_labels(filters, selected_seller, selected_customer, selected_dress)

        def _jalali_week_bounds(base_date):
            week_start = base_date - timedelta(days=base_date.weekday())
            week_end = week_start + timedelta(days=6)
            return week_start, week_end

        def _jalali_month_bounds(base_date):
            month_start = jdatetime.date(base_date.year, base_date.month, 1)
            if base_date.month == 12:
                month_end = jdatetime.date(base_date.year + 1, 1, 1) - timedelta(days=1)
            else:
                month_end = jdatetime.date(base_date.year, base_date.month + 1, 1) - timedelta(days=1)
            return month_start, month_end

        def _jalali_year_bounds(base_date):
            year_start = jdatetime.date(base_date.year, 1, 1)
            year_end = jdatetime.date(base_date.year + 1, 1, 1) - timedelta(days=1)
            return year_start, year_end

        today_jalali = jdatetime.date.fromgregorian(date=today)
        this_week_start, this_week_end = _jalali_week_bounds(today_jalali)
        this_month_start, this_month_end = _jalali_month_bounds(today_jalali)
        this_year_start, this_year_end = _jalali_year_bounds(today_jalali)

        if today_jalali.month == 12:
            next_month_start = jdatetime.date(today_jalali.year + 1, 1, 1)
            next_month_end = jdatetime.date(today_jalali.year + 1, 2, 1) - timedelta(days=1)
        else:
            next_month_start = jdatetime.date(today_jalali.year, today_jalali.month + 1, 1)
            next_month_end = jdatetime.date(today_jalali.year, today_jalali.month + 2, 1) - timedelta(days=1) if today_jalali.month < 11 else jdatetime.date(today_jalali.year + 1, 1, 1) - timedelta(days=1)

        if today_jalali.month > 3:
            three_month_start = jdatetime.date(today_jalali.year, today_jalali.month - 3, 1)
        else:
            three_month_start = jdatetime.date(today_jalali.year - 1, today_jalali.month + 9, 1)

        context['date_presets'] = [
            {'label': 'امروز', 'url': self._build_filter_url({'date_from': today.strftime('%Y-%m-%d'), 'date_to': today.strftime('%Y-%m-%d')})},
            {'label': 'این هفته', 'url': self._build_filter_url({'date_from': this_week_start.togregorian().strftime('%Y-%m-%d'), 'date_to': this_week_end.togregorian().strftime('%Y-%m-%d')})},
            {'label': 'هفته آینده', 'url': self._build_filter_url({'date_from': (this_week_start + timedelta(days=7)).togregorian().strftime('%Y-%m-%d'), 'date_to': (this_week_end + timedelta(days=7)).togregorian().strftime('%Y-%m-%d')})},
            {'label': 'این ماه', 'url': self._build_filter_url({'date_from': this_month_start.togregorian().strftime('%Y-%m-%d'), 'date_to': this_month_end.togregorian().strftime('%Y-%m-%d')})},
            {'label': 'ماه آینده', 'url': self._build_filter_url({'date_from': next_month_start.togregorian().strftime('%Y-%m-%d'), 'date_to': next_month_end.togregorian().strftime('%Y-%m-%d')})},
            {'label': '۳ ماه اخیر', 'url': self._build_filter_url({'date_from': three_month_start.togregorian().strftime('%Y-%m-%d'), 'date_to': today.strftime('%Y-%m-%d')})},
            {'label': 'سال جاری', 'url': self._build_filter_url({'date_from': this_year_start.togregorian().strftime('%Y-%m-%d'), 'date_to': this_year_end.togregorian().strftime('%Y-%m-%d')})},
        ]
        context['reset_url'] = self.request.path
        context['sellers'] = User.objects.filter(is_active=True).order_by('username')
        context['customers'] = Customer.objects.order_by('bride_last_name', 'bride_first_name')[:200]
        context['dresses'] = Dress.objects.order_by('code')[:200]
        context['statuses'] = ReservationStatus.CHOICES
        context['payment_methods'] = PaymentMethod.CHOICES
        context['report_tabs'] = [
            {'label': 'عملیاتی', 'href': '#operational'},
            {'label': 'مالی', 'href': '#financial'},
            {'label': 'مشتری', 'href': '#customers'},
            {'label': 'ریسک', 'href': '#risk'},
        ]
        top_customer_label = '—'
        if top_customer:
            top_customer_label = f"{top_customer.get('customer__bride_first_name', '')} {top_customer.get('customer__bride_last_name', '')}".strip() or '—'

        is_current_month_to_date = (date_from == current_month_start and date_to == today)
        period_title = 'این ماه' if is_current_month_to_date else 'درآمد بازه انتخابی'

        context['summary'] = {
            'month_revenue': month_revenue,
            'active_reservations': active_reservations,
            'cancelled_reservations': cancelled_reservations,
            'refund_cases': refund_cases,
            'refund_amount': refund_amount,
            'damage_penalty_total': damage_penalty_total,
            'top_dress_label': f"{top_dress['dress__code']}" if top_dress else '—',
            'top_customer_label': top_customer_label,
            'period_title': period_title,
            'period_label': f"{self._format_jalali_date(date_from)} تا {self._format_jalali_date(date_to)}",
        }
        context['financial_rows'] = [
            {'label': 'درآمد خالص از رزروها', 'value': month_revenue, 'hint': 'جمع مبلغ نهایی رزروها'},
            {'label': 'درآمد جریمه لغو', 'value': cancellation_fee_income, 'hint': 'تراکنش‌های ثبت‌شده جریمه لغو'},
            {'label': 'درآمد خسارت', 'value': damage_income, 'hint': 'دریافت خسارت از رزروها'},
            {'label': 'مجموع دریافتی‌ها', 'value': cash_inflow, 'hint': 'بیعانه + پرداخت‌های تکمیلی'},
            {'label': 'بازپرداخت‌ها', 'value': refund_total, 'hint': 'مبلغ بازپرداخت ثبت‌شده'},
            {'label': 'مانده پرداخت‌ها', 'value': outstanding_total, 'hint': 'باقی‌مانده رزروهای فعال'},
            {'label': 'هزینه‌های جانبی', 'value': additional_fee_total, 'hint': 'جمع هزینه‌های جانبی رزروها'},
            {'label': 'تخفیف و تعدیل', 'value': adjustment_total, 'hint': 'تخفیف/تعدیل ثبت‌شده'},
        ]
        context['repeat_customers'] = repeat_customers
        context['top_dresses'] = top_dresses
        context['employee_rows'] = seller_rows
        context['damage_records'] = damage_records
        context['cancellation_records'] = cancellation_records
        context['trend_labels'] = labels
        context['trend_revenue'] = revenue_series
        context['trend_reservations'] = reservation_series
        context['analysis'] = {
            'cancellation_rate': round((cancelled_reservations / reservation_qs.count() * 100) if reservation_qs.count() else 0, 1),
            'cancellation_total': cancellation_total,
            'damage_total_amount': damage_total_amount,
            'cancellation_trend': cancellation_trend,
            'damage_trend': damage_trend,
            'avg_refund': round((refund_amount / refund_cases) if refund_cases else 0, 0),
            'avg_penalty': round((sum(record.penalty_amount or 0 for record in cancellation_records) / cancellation_total) if cancellation_total else 0, 0),
        }
        return context


@login_required
def export_reports_excel(request):
    report_type = (request.GET.get('report_type') or '').strip().lower()
    view = ReportsIndexView()
    view.setup(request)
    context = view.get_context_data()

    wb = Workbook()
    wb.remove(wb.active)

    def _build_single_report_sheet(title, headers, rows, metadata=None):
        sheet = wb.create_sheet(title=title)
        _append_excel_rows(sheet, title, headers, rows, metadata=metadata)
        _apply_excel_sheet_style(sheet, title, header_row=5, metadata_count=(len(metadata or []) + 1))
        return sheet

    if report_type:
        if report_type == 'details':
            headers = [
                'شناسه رزرو',
                'شماره قرارداد',
                'مشتری',
                'لباس',
                'تاریخ شروع',
                'تاریخ پایان',
                'روزهای اجاره',
                'وضعیت رزرو',
                'وضعیت پرداخت',
                'روش پرداخت',
                'قیمت روزانه',
                'مبلغ تخفیف',
                'مبلغ نهایی',
                'بیعانه',
                'باقی‌مانده',
                'هزینه‌های جانبی',
                'خسارت',
                'جریمه لغو',
                'بازپرداخت',
                'کد رهگیری پرداخت',
            ]
            rows = []
            for reservation in Reservation.objects.filter(is_deleted=False).select_related('customer', 'dress', 'created_by'):
                customer = getattr(reservation, 'customer', None)
                dress = getattr(reservation, 'dress', None)
                rows.append([
                    reservation.pk,
                    _clean_text(getattr(reservation, 'contract_number', '') or '-'),
                    _clean_text(str(customer) if customer else '-'),
                    _clean_text(getattr(dress, 'code', '') or '-'),
                    _format_export_date(getattr(reservation, 'start_date', '')),
                    _format_export_date(getattr(reservation, 'end_date', '')),
                    getattr(reservation, 'rental_days', 0),
                    _translate_export_value(getattr(reservation, 'status', '')),
                    _translate_export_value(getattr(reservation, 'payment_status', '')),
                    _translate_export_value(getattr(reservation, 'payment_method', '')),
                    getattr(dress, 'daily_rent_price', 0) if dress else 0,
                    reservation.discount_amount or 0,
                    reservation.final_price or 0,
                    reservation.deposit_amount or 0,
                    reservation.remaining_amount or 0,
                    reservation.total_additional_fees(),
                    reservation.damage_amount or 0,
                    reservation.cancellation_fee or 0,
                    reservation.refunded_amount or 0,
                    _sanitize_tracking_code(getattr(reservation, 'payment_tracking_code', '')),
                ])
            _build_single_report_sheet(
                'جزئیات رزروها',
                headers,
                rows,
                metadata=[
                    ('تاریخ تولید', timezone.localtime(timezone.now()).strftime('%Y/%m/%d %H:%M')),
                    ('بازه زمانی', f"{context['summary']['period_label'] or '-'}"),
                ],
            )
        elif report_type == 'top_customers':
            headers = ['مشتری', 'تعداد رزرو', 'درآمد کل', 'آخرین تاریخ رزرو']
            rows = []
            for row in context.get('repeat_customers', []):
                rows.append([
                    _clean_text(f"{row.get('customer__bride_first_name') or ''} {row.get('customer__bride_last_name') or ''}".strip() or '-'),
                    row.get('reservation_count') or 0,
                    row.get('total_revenue') or 0,
                    _format_export_date(row.get('last_reservation_date')),
                ])
            _build_single_report_sheet(
                'مشتریان پرتکرار',
                headers,
                rows,
                metadata=[
                    ('تاریخ تولید', timezone.localtime(timezone.now()).strftime('%Y/%m/%d %H:%M')),
                    ('بازه زمانی', f"{context['summary']['period_label'] or '-'}"),
                ],
            )
        elif report_type == 'top_products':
            headers = ['کد لباس', 'رتبه', 'تعداد رزرو', 'درآمد کل']
            rows = []
            for index, row in enumerate(context.get('top_dresses', []), start=1):
                rows.append([
                    _clean_text(row.get('dress__code') or '-'),
                    index,
                    row.get('reservation_count') or 0,
                    row.get('total_revenue') or 0,
                ])
            _build_single_report_sheet(
                'لباس‌های پرفروش',
                headers,
                rows,
                metadata=[
                    ('تاریخ تولید', timezone.localtime(timezone.now()).strftime('%Y/%m/%d %H:%M')),
                    ('بازه زمانی', f"{context['summary']['period_label'] or '-'}"),
                ],
            )
        elif report_type == 'issues':
            headers = ['نوع', 'شناسه رزرو', 'مبلغ', 'بازپرداخت', 'تاریخ']
            rows = []
            for record in context.get('cancellation_records', []):
                rows.append(['لغو', record.reservation_id, record.penalty_amount or 0, record.refund_amount or 0, _format_export_date(getattr(record, 'cancelled_at', None))])
            for record in context.get('damage_records', []):
                rows.append(['خسارت', record.reservation_id, record.amount or 0, 0, _format_export_date(getattr(record, 'created_at', None))])
            _build_single_report_sheet(
                'لغو، بازپرداخت و خسارت',
                headers,
                rows,
                metadata=[
                    ('تاریخ تولید', timezone.localtime(timezone.now()).strftime('%Y/%m/%d %H:%M')),
                    ('بازه زمانی', f"{context['summary']['period_label'] or '-'}"),
                ],
            )
        else:
            report_type = ''

    if not report_type:
        ws = wb.create_sheet(title='خلاصه مدیریتی')
        summary_rows = [
            ('بازه زمانی', f"{context['summary']['period_label']}"),
            ('درآمد بازه انتخابی', context['summary']['month_revenue']),
            ('تعداد رزروها', context['summary']['active_reservations'] + context['summary']['cancelled_reservations']),
            ('رزروهای فعال', context['summary']['active_reservations']),
            ('رزروهای لغو شده', context['summary']['cancelled_reservations']),
            ('تعداد بازپرداخت', context['summary']['refund_cases']),
            ('مبلغ بازپرداخت', context['summary']['refund_amount']),
            ('جمع خسارت و جریمه', context['summary']['damage_penalty_total']),
            ('پرفروش‌ترین لباس', context['summary']['top_dress_label']),
            ('پرمشتری‌ترین مشتری', context['summary']['top_customer_label']),
            ('نرخ لغو', f"{context['analysis']['cancellation_rate']}٪"),
            ('میانگین بازپرداخت', context['analysis']['avg_refund']),
            ('میانگین جریمه', context['analysis']['avg_penalty']),
        ]
        _append_excel_rows(
            ws,
            'خلاصه مدیریتی',
            ['متریک', 'ارزش'],
            [(label, value) for label, value in summary_rows],
            metadata=[
                ('تاریخ تولید', timezone.localtime(timezone.now()).strftime('%Y/%m/%d %H:%M')),
                ('بازه زمانی', f"{context['summary']['period_label'] or '-'}"),
            ],
        )
        _apply_excel_sheet_style(ws, 'خلاصه مدیریتی', header_row=5)

        def _append_sheet(title, headers, rows):
            sheet = wb.create_sheet(title=title)
            _append_excel_rows(sheet, title, headers, rows, metadata=[
                ('تاریخ تولید', timezone.localtime(timezone.now()).strftime('%Y/%m/%d %H:%M')),
                ('بازه زمانی', f"{context['summary']['period_label'] or '-'}"),
            ])
            _apply_excel_sheet_style(sheet, title, header_row=5)
            return sheet

        reservation_rows = []
        for reservation in Reservation.objects.filter(is_deleted=False).select_related('customer', 'dress', 'created_by'):
            customer = getattr(reservation, 'customer', None)
            dress = getattr(reservation, 'dress', None)
            reservation_rows.append([
                reservation.pk,
                _clean_text(getattr(reservation, 'contract_number', '') or '-'),
                _clean_text(str(customer) if customer else '-'),
                _clean_text(getattr(dress, 'code', '') or '-'),
                _format_export_date(getattr(reservation, 'start_date', '')),
                _format_export_date(getattr(reservation, 'end_date', '')),
                getattr(reservation, 'rental_days', 0),
                _translate_export_value(getattr(reservation, 'status', '')),
                _translate_export_value(getattr(reservation, 'payment_status', '')),
                _translate_export_value(getattr(reservation, 'payment_method', '')),
                getattr(dress, 'daily_rent_price', 0) if dress else 0,
                reservation.discount_amount or 0,
                reservation.final_price or 0,
                reservation.deposit_amount or 0,
                reservation.remaining_amount or 0,
                reservation.total_additional_fees(),
                reservation.damage_amount or 0,
                reservation.cancellation_fee or 0,
                reservation.refunded_amount or 0,
                _sanitize_tracking_code(getattr(reservation, 'payment_tracking_code', '')),
            ])

        _append_sheet(
            'جزئیات رزروها',
            [
                'شناسه رزرو',
                'شماره قرارداد',
                'مشتری',
                'لباس',
                'تاریخ شروع',
                'تاریخ پایان',
                'روزهای اجاره',
                'وضعیت رزرو',
                'وضعیت پرداخت',
                'روش پرداخت',
                'قیمت روزانه',
                'مبلغ تخفیف',
                'مبلغ نهایی',
                'بیعانه',
                'باقی‌مانده',
                'هزینه‌های جانبی',
                'خسارت',
                'جریمه لغو',
                'بازپرداخت',
                'کد رهگیری پرداخت',
            ],
            reservation_rows,
        )

        trend_labels = context.get('trend_labels', [])
        trend_revenue = context.get('trend_revenue', [])
        trend_reservations = context.get('trend_reservations', [])
        daily_rows = [
            [label, revenue, reservation_count]
            for label, revenue, reservation_count in zip(trend_labels, trend_revenue, trend_reservations)
        ]
        _append_sheet('روند روزانه', ['روز', 'درآمد', 'تعداد رزرو'], daily_rows)

        repeat_customer_rows = []
        for row in context.get('repeat_customers', []):
            repeat_customer_rows.append([
                _clean_text(f"{row.get('customer__bride_first_name') or ''} {row.get('customer__bride_last_name') or ''}".strip() or '-'),
                row.get('reservation_count') or 0,
                row.get('total_revenue') or 0,
            ])
        _append_sheet('مشتریان پرتکرار', ['مشتری', 'تعداد رزرو', 'درآمد کل'], repeat_customer_rows)

        top_dress_rows = []
        for row in context.get('top_dresses', []):
            top_dress_rows.append([
                _clean_text(row.get('dress__code') or '-'),
                row.get('reservation_count') or 0,
                row.get('total_revenue') or 0,
            ])
        _append_sheet('لباس‌های پرفروش', ['کد لباس', 'تعداد رزرو', 'درآمد کل'], top_dress_rows)

        seller_rows = []
        for row in context.get('employee_rows', []):
            seller_rows.append([
                _user_label(User.objects.filter(pk=row.get('created_by__id')).first()) if row.get('created_by__id') else '-',
                row.get('reservation_count') or 0,
                row.get('total_revenue') or 0,
                row.get('cancelled_count') or 0,
                row.get('damage_total') or 0,
            ])
        _append_sheet('عملکرد فروشندگان', ['فروشنده', 'تعداد رزرو', 'درآمد کل', 'لغو شده', 'خسارت'], seller_rows)

        analysis_rows = []
        for record in context.get('cancellation_records', []):
            analysis_rows.append([
                'لغو',
                record.reservation_id,
                record.penalty_amount or 0,
                record.refund_amount or 0,
                _format_export_date(getattr(record, 'cancelled_at', None)),
            ])
        for record in context.get('damage_records', []):
            analysis_rows.append([
                'خسارت',
                record.reservation_id,
                record.amount or 0,
                0,
                _format_export_date(getattr(record, 'created_at', None)),
            ])
        _append_sheet('لغو، بازپرداخت و خسارت', ['نوع', 'شناسه رزرو', 'مبلغ', 'بازپرداخت', 'تاریخ'], analysis_rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reports_export.xlsx"'
    return response
